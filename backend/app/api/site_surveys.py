from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Body, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import List, Optional
from datetime import datetime
import io
import os
import uuid
import zipfile
import csv

from app.core.database import get_db
from app.api.auth import get_current_user
from app.models.user import User
from app.models.site import Site
from app.models.survey import SiteSurvey, SiteSurveyPhoto
from app.models.work_order import AuditEvent
from app.schemas.site_survey import (
    SiteSurveyCreate, SiteSurveyUpdate,
    SiteSurveyResponse, SiteSurveyPhotoResponse,
    SiteSurveySummary,
)
from app.utils.file_handler import save_uploaded_file, calculate_file_hash, extract_exif, compress_image, add_text_watermark_inline
from app.utils.timezone import to_utc_iso
from app.utils.archive_pdf import (
    build_pdf_styles,
    create_pdf_cell,
    localized_text,
    normalize_locale,
    register_pdf_fonts,
)


router = APIRouter()


def _require_admin(current_user: User) -> None:
    if getattr(current_user, "role", None) not in ("admin", "manager"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="需要管理员/经理权限")


def _can_edit(db: Session, current_user: User, survey: SiteSurvey) -> bool:
    linked_work_order = None
    if getattr(survey, 'work_order_id', None):
        from app.models.work_order import WorkOrder, WorkOrderTypeEnum, WorkOrderStatusEnum

        linked_work_order = db.query(WorkOrder).filter(WorkOrder.id == survey.work_order_id).first()
        if linked_work_order and linked_work_order.status == WorkOrderStatusEnum.VOIDED:
            return False

    role = getattr(current_user, 'role', None)
    if role in ("admin", "manager"):
        return True
    if role == "surveyor" and getattr(survey, 'work_order_id', None):
        from app.models.work_order import WorkOrderTypeEnum, WorkOrderStatusEnum

        wo = linked_work_order
        if not wo:
            return False
        if wo.assigned_to != current_user.id:
            return False
        if wo.type != WorkOrderTypeEnum.SITE_SURVEY:
            return False
        if wo.status not in (WorkOrderStatusEnum.ACTIVE, WorkOrderStatusEnum.REJECTED):
            return False
        return True
    return False


def _audit(
    db: Session,
    resource_id: str,
    action: str,
    operator_id: int,
    comments: Optional[str] = None,
    details: Optional[dict] = None,
):
    evt = AuditEvent(
        id=uuid.uuid4().hex,
        resource_type="site_survey",
        resource_id=resource_id,
        action=action,
        operator_id=operator_id,
        comments=comments,
        details=details or {},
    )
    db.add(evt)


def _jsonify_val(v):
    if isinstance(v, datetime):
        try:
            return to_utc_iso(v)
        except Exception:
            return str(v)
    return v

def _jsonify_dict(d: dict) -> dict:
    try:
        return {k: _jsonify_val(v) for k, v in d.items()}
    except Exception:
        return {}

# Human-readable labels for common fields
FIELD_LABELS = {
    "survey_date": "勘察日期",
    "surveyor_name": "勘察人",
    "surveyor_phone": "勘察人电话",
    "feasibility": "结论",
    "address": "地址",
    "latitude": "纬度",
    "longitude": "经度",
    "risks": "风险",
    "recommendations": "建议",
}

def _fmt_val(v):
    if isinstance(v, datetime):
        try:
            return v.strftime('%Y-%m-%d %H:%M')
        except Exception:
            return to_utc_iso(v)
    if isinstance(v, float):
        return round(v, 6)
    if v is None:
        return "-"
    return v

def _values_different(a, b) -> bool:
    # Normalize datetimes
    if isinstance(a, datetime) and isinstance(b, datetime):
        return a.replace(microsecond=0) != b.replace(microsecond=0)
    # Floats tolerance
    if isinstance(a, float) and isinstance(b, float):
        return abs(a - b) > 1e-9
    return a != b


@router.post("/", response_model=SiteSurveyResponse)
async def create_survey(
    payload: SiteSurveyCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    _require_admin(current_user)

    site = db.query(Site).filter(Site.id == payload.site_id).first()
    if not site:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="站点不存在")

    survey = SiteSurvey(
        id=uuid.uuid4().hex,
        created_by=current_user.id,
        **payload.dict(exclude_unset=True),
    )
    db.add(survey)
    db.commit()
    # audit
    try:
        _audit(db, survey.id, "create", current_user.id, details={"site_id": survey.site_id})
        db.commit()
    except Exception:
        db.rollback()
    db.refresh(survey)
    return survey


@router.get("/import-template")
async def download_import_template():
    """下载Excel导入模板

    注意：此静态路径必须在 `/{survey_id}` 动态路径之前注册，
    否则会被当作 survey_id 捕获导致 404。
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "SiteSurveys"
    headers = [
        "site_code",
        "survey_date(YYYY-MM-DD)",
        "surveyor_name",
        "surveyor_phone",
        "feasibility(feasible|conditionally_feasible|infeasible)",
        "latitude",
        "longitude",
        "address",
        "power_available(true/false)",
        "fiber_available(true/false)",
        "microwave_los(true/false)",
        "los_azimuth_deg",
        "los_distance_km",
        "risks",
        "recommendations",
    ]
    ws.append(headers)
    # one sample row
    ws.append([
        "BJ-001", "2025-10-23", "张三", "13800000000", "feasible",
        39.9, 116.3, "北京市朝阳区XX路", True, True, False, None, None, "周边学校", "建议使用光纤回传",
    ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=site_survey_import_template.xlsx"}
    )


@router.get("/page")
async def page_surveys(
    page: int = 1,
    page_size: int = 20,
    keyword: Optional[str] = None,
    site_id: Optional[int] = None,
    feasibility: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    sort_by: Optional[str] = "created_at",  # created_at|survey_date|site_code
    sort_dir: Optional[str] = "desc",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(SiteSurvey).join(Site, Site.id == SiteSurvey.site_id)
    if site_id:
        query = query.filter(SiteSurvey.site_id == site_id)
    if feasibility:
        query = query.filter(SiteSurvey.feasibility == feasibility)
    if date_from:
        query = query.filter(SiteSurvey.survey_date >= date_from)
    if date_to:
        query = query.filter(SiteSurvey.survey_date <= date_to)
    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(or_(
            Site.site_name.like(kw),
            Site.site_code.like(kw),
            SiteSurvey.surveyor_name.like(kw),
            SiteSurvey.feasibility.like(kw)
        ))

    # Total
    total = query.with_entities(func.count(SiteSurvey.id)).scalar() or 0

    # Sorting
    sort_expr = SiteSurvey.created_at.desc()
    if sort_by == "survey_date":
        sort_expr = SiteSurvey.survey_date.desc() if sort_dir == "desc" else SiteSurvey.survey_date.asc()
    elif sort_by == "site_code":
        sort_expr = Site.site_code.desc() if sort_dir == "desc" else Site.site_code.asc()
    else:
        sort_expr = SiteSurvey.created_at.desc() if sort_dir == "desc" else SiteSurvey.created_at.asc()

    rows = (query.order_by(sort_expr)
                 .offset((page - 1) * page_size)
                 .limit(page_size)
                 .all())

    items: List[SiteSurveySummary] = []
    for s in rows:
        items.append(
            SiteSurveySummary(
                id=s.id,
                site_id=s.site_id,
                site_name=s.site.site_name if s.site else None,
                site_code=s.site.site_code if s.site else None,
                survey_date=s.survey_date,
                surveyor_name=s.surveyor_name,
                feasibility=s.feasibility,
                created_at=s.created_at,
            )
        )
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """批量导入勘察数据（Excel）

    注意：置于 `/{survey_id}` 动态路径之前，避免被误匹配。
    """
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限上传该勘察照片")
    from openpyxl import load_workbook

    content = await file.read()
    stream = io.BytesIO(content)
    wb = load_workbook(stream)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = [cell.value for cell in header_row[0:]]

    success = 0
    failed = 0
    errors: List[dict] = []

    # Build simple site_code -> id map
    sites = db.query(Site).all()
    site_by_code = {s.site_code: s for s in sites}

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            site_code = row[0]
            survey_date_s = row[1]
            surveyor_name = row[2]
            surveyor_phone = row[3]
            feasibility = row[4]

            if not site_code or not survey_date_s or not surveyor_name or not feasibility:
                raise ValueError("必填列缺失: site_code/survey_date/surveyor_name/feasibility")

            site_obj = site_by_code.get(str(site_code))
            if not site_obj:
                raise ValueError(f"站点编码不存在: {site_code}")

            # Parse date
            if isinstance(survey_date_s, datetime):
                survey_date = survey_date_s
            else:
                survey_date = datetime.fromisoformat(str(survey_date_s))

            survey = SiteSurvey(
                id=uuid.uuid4().hex,
                site_id=site_obj.id,
                survey_date=survey_date,
                surveyor_name=str(surveyor_name),
                surveyor_phone=str(surveyor_phone) if surveyor_phone else None,
                feasibility=str(feasibility),
                latitude=float(row[5]) if row[5] is not None else None,
                longitude=float(row[6]) if row[6] is not None else None,
                address=str(row[7]) if row[7] is not None else None,
                power_available=bool(row[8]) if row[8] is not None else None,
                fiber_available=bool(row[9]) if row[9] is not None else None,
                microwave_los=bool(row[10]) if row[10] is not None else None,
                los_azimuth_deg=float(row[11]) if row[11] is not None else None,
                los_distance_km=float(row[12]) if row[12] is not None else None,
                risks=str(row[13]) if row[13] is not None else None,
                recommendations=str(row[14]) if row[14] is not None else None,
                created_by=current_user.id,
            )
            db.add(survey)
            success += 1
        except Exception as e:
            failed += 1
            row_dict = {str(header[idx] or idx): row[idx] for idx in range(min(len(header), len(row)))}
            errors.append({"row_index": i, "row_values": row_dict, "error": str(e)})

    db.commit()
    # audit each created survey (optional lightweight)
    try:
        # 为避免逐条查询，简单记录一次导入汇总
        _audit(db, resource_id="bulk", action="import", operator_id=current_user.id,
               details={"success": success, "failed": failed})
        db.commit()
    except Exception:
        db.rollback()

    return {"success": success, "failed": failed, "errors": errors}

@router.get("/", response_model=List[SiteSurveySummary])
async def list_surveys(
    skip: int = 0,
    limit: int = 50,
    site_id: Optional[int] = None,
    surveyor_name: Optional[str] = None,
    feasibility: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(SiteSurvey)
    if site_id:
        query = query.filter(SiteSurvey.site_id == site_id)
    if surveyor_name:
        query = query.filter(SiteSurvey.surveyor_name.like(f"%{surveyor_name}%"))
    if feasibility:
        query = query.filter(SiteSurvey.feasibility == feasibility)
    if date_from:
        query = query.filter(SiteSurvey.survey_date >= date_from)
    if date_to:
        query = query.filter(SiteSurvey.survey_date <= date_to)

    rows = query.order_by(SiteSurvey.created_at.desc()).offset(skip).limit(limit).all()

    summaries: List[SiteSurveySummary] = []
    for s in rows:
        summaries.append(
            SiteSurveySummary(
                id=s.id,
                site_id=s.site_id,
                site_name=s.site.site_name if s.site else None,
                site_code=s.site.site_code if s.site else None,
                survey_date=s.survey_date,
                surveyor_name=s.surveyor_name,
                feasibility=s.feasibility,
                created_at=s.created_at,
            )
        )
    return summaries


 


@router.get("/{survey_id}", response_model=SiteSurveyResponse)
async def get_survey(
    survey_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="勘察记录不存在")

    # Ordered photos
    photos = (
        db.query(SiteSurveyPhoto)
        .filter(SiteSurveyPhoto.survey_id == survey_id)
        .order_by(SiteSurveyPhoto.sort_order.asc().nulls_last(), SiteSurveyPhoto.created_at.asc())
        .all()
    )

    # Build response explicitly to包含 site_name/site_code
    from app.schemas.site_survey import SiteSurveyPhotoResponse  # local import to avoid cycle
    photo_items = [SiteSurveyPhotoResponse.model_validate(p) for p in photos]

    return {
        "id": survey.id,
        "site_id": survey.site_id,
        "site_name": survey.site.site_name if survey.site else None,
        "site_code": survey.site.site_code if survey.site else None,
        "survey_date": survey.survey_date,
        "surveyor_name": survey.surveyor_name,
        "surveyor_phone": survey.surveyor_phone,
        "latitude": survey.latitude,
        "longitude": survey.longitude,
        "address": survey.address,
        "gps_accuracy": survey.gps_accuracy,
        "site_type": survey.site_type,
        "tower_type": survey.tower_type,
        "available_height_m": survey.available_height_m,
        "load_capacity_kg": survey.load_capacity_kg,
        "power_available": survey.power_available,
        "power_distance_m": survey.power_distance_m,
        "power_capacity_kw": survey.power_capacity_kw,
        "earthing_feasible": survey.earthing_feasible,
        "fiber_available": survey.fiber_available,
        "fiber_distance_m": survey.fiber_distance_m,
        "duct_notes": survey.duct_notes,
        "microwave_los": survey.microwave_los,
        "los_azimuth_deg": survey.los_azimuth_deg,
        "los_distance_km": survey.los_distance_km,
        "sensitive_points": survey.sensitive_points,
        "safety_notes": survey.safety_notes,
        "permits_constraints": survey.permits_constraints,
        "owner_name": survey.owner_name,
        "owner_phone": survey.owner_phone,
        "access_time_window": survey.access_time_window,
        "entry_constraints": survey.entry_constraints,
        "feasibility": survey.feasibility,
        "risks": survey.risks,
        "recommendations": survey.recommendations,
        "extra_data": survey.extra_data,
        "created_by": survey.created_by,
        "created_at": survey.created_at,
        "updated_at": survey.updated_at,
        "photos": photo_items,
    }


@router.put("/{survey_id}", response_model=SiteSurveyResponse)
async def update_survey(
    survey_id: str,
    payload: SiteSurveyUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="勘察记录不存在")
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限编辑该勘察记录")
    incoming = payload.dict(exclude_unset=True)
    # build diff（仅记录值真正发生变化的字段）
    before = {}
    after = {}
    for k, v in incoming.items():
        old = getattr(survey, k)
        if _values_different(old, v):
            before[k] = old
            after[k] = v
        # 仍然赋值，确保更新
        setattr(survey, k, v)
    survey.updated_at = datetime.utcnow()
    db.commit()
    # audit（仅当有真实变更时记录）
    try:
        if after:
            changed = []
            for k in after.keys():
                changed.append({
                    "field": k,
                    "label": FIELD_LABELS.get(k, k),
                    "before": _fmt_val(before.get(k)),
                    "after": _fmt_val(after.get(k)),
                })
            _audit(
                db, survey.id, "update", current_user.id,
                details={
                    "before": _jsonify_dict(before),
                    "after": _jsonify_dict(after),
                    "changed": changed,
                    "changed_keys": list(after.keys()),
                }
            )
            db.commit()
    except Exception:
        db.rollback()
    db.refresh(survey)
    return survey


@router.delete("/{survey_id}")
async def delete_survey(
    survey_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="勘察记录不存在")
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限删除该勘察记录")

    # 删除文件
    for p in list(survey.photos or []):
        try:
            if p.file_path and os.path.exists(p.file_path):
                os.remove(p.file_path)
        except Exception:
            pass
    db.delete(survey)
    db.commit()
    # audit
    try:
        _audit(db, survey_id, "delete", current_user.id, details={"site_id": survey.site_id})
        db.commit()
    except Exception:
        db.rollback()
    return {"success": True}


@router.post("/{survey_id}/photos", response_model=SiteSurveyPhotoResponse)
async def upload_photo(
    survey_id: str,
    file: UploadFile = File(...),
    category: Optional[str] = Form(None),
    latitude: Optional[float] = Form(None),
    longitude: Optional[float] = Form(None),
    gps_accuracy: Optional[float] = Form(None),
    address: Optional[str] = Form(None),
    taken_at: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="勘察记录不存在")
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限上传该勘察照片")

    stored_path = await save_uploaded_file(file, category="site_surveys", sub_folder=survey_id)
    file_hash = calculate_file_hash(stored_path)

    # 先从原始文件提取 EXIF，再执行压缩/水印，避免处理过程中丢失元数据
    is_image = (file.content_type or '').startswith('image/')
    raw_exif = None
    if is_image:
        try:
            raw_exif = extract_exif(stored_path)
        except Exception:
            raw_exif = None
        try:
            # 压缩
            stored_path = await compress_image(stored_path)
            # 添加水印（覆盖保存）
            wm_text = f"Survey {survey_id} | {current_user.username} | {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
            add_text_watermark_inline(stored_path, wm_text)
        except Exception:
            pass

    taken_dt: Optional[datetime] = None
    if taken_at:
        try:
            taken_dt = datetime.fromisoformat(taken_at.replace("Z", "+00:00"))
        except Exception:
            taken_dt = None

    # Auto taken_at from EXIF if not provided
    if not taken_dt and isinstance(raw_exif, dict):
        dt_str = raw_exif.get('DateTimeOriginal') or raw_exif.get('DateTime')
        if isinstance(dt_str, str):
            try:
                # EXIF format "YYYY:MM:DD HH:MM:SS"
                dt_fmt = dt_str.replace(':', '-', 2)
                taken_dt = datetime.fromisoformat(dt_fmt)
            except Exception:
                taken_dt = None

    # Determine next sort_order
    max_order = db.query(func.max(SiteSurveyPhoto.sort_order)).filter(SiteSurveyPhoto.survey_id == survey_id).scalar()
    next_order = (max_order or 0) + 1

    photo = SiteSurveyPhoto(
        id=uuid.uuid4().hex,
        survey_id=survey_id,
        original_name=file.filename,
        file_path=stored_path,
        file_size=None,
        mime_type=file.content_type,
        category=category,
        sort_order=next_order,
        latitude=latitude,
        longitude=longitude,
        gps_accuracy=gps_accuracy,
        address=address,
        taken_at=taken_dt,
        hash_value=file_hash,
        uploaded_by=current_user.id,
    )
    try:
        if os.path.exists(stored_path):
            photo.file_size = os.path.getsize(stored_path)
    except Exception:
        pass

    # Persist EXIF json if available
    if raw_exif:
        try:
            photo.exif = raw_exif
        except Exception:
            pass
    db.add(photo)
    db.commit()
    # audit
    try:
        _audit(db, survey_id, "photo_upload", current_user.id, details={
            "photo_id": photo.id,
            "category": category,
            "mime_type": file.content_type,
            "original_name": file.filename,
        })
        db.commit()
    except Exception:
        db.rollback()
    db.refresh(photo)
    return photo


@router.get("/{survey_id}/photos", response_model=List[SiteSurveyPhotoResponse])
async def list_photos(
    survey_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    photos = (
        db.query(SiteSurveyPhoto)
        .filter(SiteSurveyPhoto.survey_id == survey_id)
        .order_by(SiteSurveyPhoto.sort_order.asc().nulls_last(), SiteSurveyPhoto.created_at.asc())
        .all()
    )
    return photos


@router.delete("/photos/{photo_id}")
async def delete_photo(
    photo_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 校验删除权限
    p = db.query(SiteSurveyPhoto).filter(SiteSurveyPhoto.id == photo_id).first()
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="照片不存在")
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == p.survey_id).first()
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权限删除该照片")
    try:
        if p.file_path and os.path.exists(p.file_path):
            os.remove(p.file_path)
    except Exception:
        pass
    # capture for audit
    details = {"photo_id": p.id, "category": p.category, "original_name": p.original_name}
    db.delete(p)
    db.commit()
    try:
        _audit(db, p.survey_id, "photo_delete", current_user.id, details=details)
        db.commit()
    except Exception:
        db.rollback()
    return {"success": True}




@router.get("/{survey_id}/export")
async def export_survey_zip(
    survey_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="勘察记录不存在")

    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # JSON summary
        # 覆盖全部勘察信息字段，保证压缩包信息完备
        summary = {
            "id": survey.id,
            "site_id": survey.site_id,
            "site_code": survey.site.site_code if survey.site else None,
            "site_name": survey.site.site_name if survey.site else None,
            "survey_date": to_utc_iso(survey.survey_date) if survey.survey_date else None,
            "surveyor_name": survey.surveyor_name,
            "surveyor_phone": survey.surveyor_phone,
            "feasibility": survey.feasibility,
            "address": survey.address,
            # Geo
            "latitude": survey.latitude, "longitude": survey.longitude, "gps_accuracy": survey.gps_accuracy,
            # Site/Structure
            "site_type": survey.site_type, "tower_type": survey.tower_type,
            "available_height_m": survey.available_height_m, "load_capacity_kg": survey.load_capacity_kg,
            # Power/Earthing
            "power_available": survey.power_available, "power_distance_m": survey.power_distance_m,
            "power_capacity_kw": survey.power_capacity_kw, "earthing_feasible": survey.earthing_feasible,
            # Transmission
            "fiber_available": survey.fiber_available, "fiber_distance_m": survey.fiber_distance_m,
            "microwave_los": survey.microwave_los, "los_azimuth_deg": survey.los_azimuth_deg,
            "los_distance_km": survey.los_distance_km,
            # Environment/Access/Owner
            "sensitive_points": survey.sensitive_points, "safety_notes": survey.safety_notes,
            "permits_constraints": survey.permits_constraints, "entry_constraints": survey.entry_constraints,
            "owner_name": survey.owner_name, "owner_phone": survey.owner_phone,
            "access_time_window": survey.access_time_window,
            # Conclusion & notes
            "risks": survey.risks, "recommendations": survey.recommendations,
        }
        zf.writestr("summary.json", __import__("json").dumps(summary, ensure_ascii=False, indent=2))

        # CSV（包含全部键值）
        csv_buf = io.StringIO()
        writer = csv.writer(csv_buf)
        writer.writerow(["key", "value"])
        for k, v in summary.items():
            writer.writerow([k, v])
        zf.writestr("summary.csv", csv_buf.getvalue())

        # Files (photos + attachments)
        photos = db.query(SiteSurveyPhoto).filter(SiteSurveyPhoto.survey_id == survey_id).all()
        # files.csv 列出所有文件元数据
        files_csv = io.StringIO()
        fw = csv.writer(files_csv)
        fw.writerow(["id","category","mime_type","file_name","file_path","file_size","taken_at","latitude","longitude"]) 
        for p in photos:
            if p.file_path and os.path.exists(p.file_path):
                # put under photos/{category}/filename
                base_name = os.path.basename(p.file_path)
                subdir = f"photos/{p.category or 'uncategorized'}"
                arcname = f"{subdir}/{base_name}"
                zf.write(p.file_path, arcname=arcname)
                fw.writerow([
                    p.id, p.category, p.mime_type, base_name, arcname,
                    (os.path.getsize(p.file_path) if os.path.exists(p.file_path) else None),
                    (to_utc_iso(p.taken_at) if p.taken_at else None), p.latitude, p.longitude
                ])
        zf.writestr("files.csv", files_csv.getvalue())

    mem_zip.seek(0)
    filename = f"site_survey_{survey_id}.zip"
    return StreamingResponse(mem_zip, media_type="application/zip", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export-batch")
async def export_batch_zip(
    keyword: Optional[str] = None,
    site_id: Optional[int] = None,
    feasibility: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    include_photos: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Build query
    query = db.query(SiteSurvey).join(Site, Site.id == SiteSurvey.site_id)
    if site_id:
        query = query.filter(SiteSurvey.site_id == site_id)
    if feasibility:
        query = query.filter(SiteSurvey.feasibility == feasibility)
    if date_from:
        query = query.filter(SiteSurvey.survey_date >= date_from)
    if date_to:
        query = query.filter(SiteSurvey.survey_date <= date_to)
    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(or_(
            Site.site_name.like(kw),
            Site.site_code.like(kw),
            SiteSurvey.surveyor_name.like(kw),
            SiteSurvey.feasibility.like(kw)
        ))

    rows = query.order_by(SiteSurvey.created_at.desc()).all()

    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        index_csv = io.StringIO()
        writer = csv.writer(index_csv)
        writer.writerow(["survey_id", "site_code", "site_name", "survey_date", "surveyor_name", "feasibility"]) 

        if not rows:
            # 写入空索引与说明文件，返回空包而非404
            zf.writestr("index.csv", index_csv.getvalue())
            zf.writestr("README.txt", "No survey records matched the export filters.")
        else:
            for s in rows:
                site_code = s.site.site_code if s.site else ''
                site_name = s.site.site_name if s.site else ''
                survey_folder = f"{site_code or 'site'}/{s.id}"

                summary = {
                    "id": s.id,
                    "site_id": s.site_id,
                    "site_code": site_code,
                    "site_name": site_name,
                    "survey_date": to_utc_iso(s.survey_date) if s.survey_date else None,
                    "surveyor_name": s.surveyor_name,
                    "surveyor_phone": s.surveyor_phone,
                    "feasibility": s.feasibility,
                    "address": s.address,
                    # Geo
                    "latitude": s.latitude, "longitude": s.longitude, "gps_accuracy": s.gps_accuracy,
                    # Site/Structure
                    "site_type": s.site_type, "tower_type": s.tower_type,
                    "available_height_m": s.available_height_m, "load_capacity_kg": s.load_capacity_kg,
                    # Power/Earthing
                    "power_available": s.power_available, "power_distance_m": s.power_distance_m,
                    "power_capacity_kw": s.power_capacity_kw, "earthing_feasible": s.earthing_feasible,
                    # Transmission
                    "fiber_available": s.fiber_available, "fiber_distance_m": s.fiber_distance_m,
                    "microwave_los": s.microwave_los, "los_azimuth_deg": s.los_azimuth_deg,
                    "los_distance_km": s.los_distance_km,
                    # Environment/Access/Owner
                    "sensitive_points": s.sensitive_points, "safety_notes": s.safety_notes,
                    "permits_constraints": s.permits_constraints, "entry_constraints": s.entry_constraints,
                    "owner_name": s.owner_name, "owner_phone": s.owner_phone,
                    "access_time_window": s.access_time_window,
                    # Conclusion & notes
                    "risks": s.risks, "recommendations": s.recommendations,
                }
                zf.writestr(f"{survey_folder}/summary.json", __import__("json").dumps(summary, ensure_ascii=False, indent=2))

                csv_buf = io.StringIO()
                w = csv.writer(csv_buf)
                for k, v in summary.items():
                    w.writerow([k, v])
                zf.writestr(f"{survey_folder}/summary.csv", csv_buf.getvalue())

                writer.writerow([
                    s.id, site_code, site_name, summary["survey_date"], s.surveyor_name, s.feasibility
                ])

                photos = db.query(SiteSurveyPhoto).filter(SiteSurveyPhoto.survey_id == s.id).all()
                # 写入文件索引 CSV
                files_csv = io.StringIO()
                fw = csv.writer(files_csv)
                fw.writerow(["id","category","mime_type","file_name","file_path","file_size","taken_at","latitude","longitude"]) 
                for p in photos:
                    base = os.path.basename(p.file_path) if p.file_path else ''
                    rel = f"{survey_folder}/photos/{p.category or 'uncategorized'}/{base}" if base else ''
                    if include_photos and p.file_path and os.path.exists(p.file_path):
                        pdir = f"{survey_folder}/photos/{p.category or 'uncategorized'}"
                        zf.write(p.file_path, arcname=f"{pdir}/{base}")
                    fw.writerow([
                        p.id, p.category, p.mime_type, base, rel,
                        (os.path.getsize(p.file_path) if p.file_path and os.path.exists(p.file_path) else None),
                        (to_utc_iso(p.taken_at) if p.taken_at else None), p.latitude, p.longitude
                    ])
                zf.writestr(f"{survey_folder}/files.csv", files_csv.getvalue())

            zf.writestr("index.csv", index_csv.getvalue())

    mem_zip.seek(0)
    return StreamingResponse(mem_zip, media_type="application/zip", headers={"Content-Disposition": "attachment; filename=site_surveys_batch.zip"})


@router.get("/{survey_id}/audit-logs")
async def get_survey_audit_logs(
    survey_id: str,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 允许所有已登录用户查看此勘察的审计历史
    logs = (
        db.query(AuditEvent)
        .filter(AuditEvent.resource_type == "site_survey", AuditEvent.resource_id.in_([survey_id, "bulk"]))
        .order_by(AuditEvent.created_at.desc())
        .offset(skip).limit(limit).all()
    )
    data = []
    for ev in logs:
        op = ev.operator
        data.append({
            "id": ev.id,
            "action": ev.action,
            "operator_id": ev.operator_id,
            "operator_name": (op.full_name or op.username) if op else None,
            "comments": ev.comments,
            "details": ev.details,
            "created_at": to_utc_iso(ev.created_at) if ev.created_at else None,
        })
    return data

@router.put("/photos/{photo_id}", response_model=SiteSurveyPhotoResponse)
async def update_photo_metadata(
    photo_id: str,
    category: Optional[str] = Form(None, alias="category"),
    taken_at: Optional[str] = Form(None, alias="taken_at"),
    sort_order: Optional[int] = Form(None, alias="sort_order"),
    # 兼容 JSON 载荷
    category_json: Optional[str] = Body(None, embed=True, alias="category"),
    taken_at_json: Optional[str] = Body(None, embed=True, alias="taken_at"),
    sort_order_json: Optional[int] = Body(None, embed=True, alias="sort_order"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # 权限：仅管理员/经理或关联工单的勘察人员可调整元数据
    p = db.query(SiteSurveyPhoto).filter(SiteSurveyPhoto.id == photo_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="照片不存在")
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == p.survey_id).first()
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=403, detail="无权限调整该照片")
    before = {"category": p.category, "taken_at": to_utc_iso(p.taken_at) if p.taken_at else None, "sort_order": p.sort_order}
    # 以 form 值优先；若缺失则使用 JSON 值
    eff_category = category if category is not None else category_json
    eff_sort = sort_order if sort_order is not None else sort_order_json
    eff_taken = taken_at if taken_at is not None else taken_at_json
    if eff_category is not None:
        p.category = eff_category
    if eff_sort is not None:
        p.sort_order = eff_sort
    if eff_taken:
        try:
            p.taken_at = datetime.fromisoformat(eff_taken.replace("Z", "+00:00"))
        except Exception:
            pass
    db.commit()
    db.refresh(p)
    after = {"category": p.category, "taken_at": to_utc_iso(p.taken_at) if p.taken_at else None, "sort_order": p.sort_order}
    try:
        # Build changed list for readability
        changed = []
        for k in after.keys():
            if before.get(k) != after.get(k):
                lbl = {"category": "分类", "taken_at": "拍摄时间", "sort_order": "显示顺序"}.get(k, k)
                changed.append({"field": k, "label": lbl, "before": before.get(k), "after": after.get(k)})
        _audit(db, p.survey_id, "photo_update", current_user.id, details={"before": before, "after": after, "photo_id": p.id, "changed": changed, "changed_keys": [c["field"] for c in changed]})
        db.commit()
    except Exception:
        db.rollback()
    return p


# 照片排序功能已废弃：如需恢复，请参考历史实现（photo_reorder）。


@router.post("/{survey_id}/photos/batch")
async def upload_photos_batch(
    survey_id: str,
    files: List[UploadFile] = File(...),
    category: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="勘察记录不存在")
    if not _can_edit(db, current_user, survey):
        raise HTTPException(status_code=403, detail="无权限上传该勘察照片")
    results = []
    for f in files:
        # Reuse single upload flow by calling within same request logic
        stored_path = await save_uploaded_file(f, category="site_surveys", sub_folder=survey_id)
        file_hash = calculate_file_hash(stored_path)
        is_image = (f.content_type or '').startswith('image/')
        # 先提取原始 EXIF，再压缩/水印
        raw_exif = None
        if is_image:
            try:
                raw_exif = extract_exif(stored_path)
            except Exception:
                raw_exif = None
            try:
                stored_path = await compress_image(stored_path)
                wm_text = f"Survey {survey_id} | {current_user.username} | {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
                add_text_watermark_inline(stored_path, wm_text)
            except Exception:
                pass
        max_order = db.query(func.max(SiteSurveyPhoto.sort_order)).filter(SiteSurveyPhoto.survey_id == survey_id).scalar()
        next_order = (max_order or 0) + 1
        p = SiteSurveyPhoto(
            id=uuid.uuid4().hex,
            survey_id=survey_id,
            original_name=f.filename,
            file_path=stored_path,
            file_size=os.path.getsize(stored_path) if os.path.exists(stored_path) else None,
            mime_type=f.content_type,
            category=category,
            sort_order=next_order,
            has_watermark=is_image,
            hash_value=file_hash,
            uploaded_by=current_user.id,
        )
        if raw_exif:
            p.exif = raw_exif
        db.add(p)
        results.append({"id": p.id, "file_path": p.file_path, "category": p.category})
    db.commit()
    try:
        _audit(db, survey_id, "photo_batch_upload", current_user.id, details={"count": len(results), "items": results[:10]})
        db.commit()
    except Exception:
        db.rollback()
    return {"success": True, "count": len(results), "items": results}


@router.delete("/{survey_id}/photos")
async def delete_photos_batch(
    survey_id: str,
    photo_ids: List[str] = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    _require_admin(current_user)
    photos = db.query(SiteSurveyPhoto).filter(SiteSurveyPhoto.survey_id == survey_id, SiteSurveyPhoto.id.in_(photo_ids)).all()
    deleted = []
    for p in photos:
        try:
            if p.file_path and os.path.exists(p.file_path):
                os.remove(p.file_path)
        except Exception:
            pass
        deleted.append({"photo_id": p.id, "category": p.category, "original_name": p.original_name})
        db.delete(p)
    db.commit()
    try:
        _audit(db, survey_id, "photo_batch_delete", current_user.id, details={"count": len(deleted), "items": [x["photo_id"] for x in deleted]})
        db.commit()
    except Exception:
        db.rollback()
    return {"success": True, "deleted": len(photos)}


# EXIF查看接口已废弃：如需恢复，请参考历史实现。


@router.get("/{survey_id}/export-pdf")
async def export_survey_pdf(
    survey_id: str,
    with_thumbs: bool = True,
    locale: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    try:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    except Exception:
        raise HTTPException(status_code=500, detail="后端缺少 reportlab 依赖，请先安装")

    survey = db.query(SiteSurvey).filter(SiteSurvey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="勘察记录不存在")

    locale_code = normalize_locale(locale)

    # Register Chinese-capable fonts
    def register_cn_fonts() -> tuple[str, str]:
        return register_pdf_fonts()

    FONT_MAIN, FONT_BOLD = register_cn_fonts()
    primary = colors.HexColor("#F56C3A")  # Orange primary
    accent_bg = colors.HexColor("#FFF6F0")  # Light orange
    text_color = colors.HexColor("#333333")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=1.8*cm
    )

    styles = build_pdf_styles(FONT_MAIN, FONT_BOLD)

    def draw_page_frame(c, d):
        c.saveState()
        # Header bar
        c.setFillColor(primary)
        c.rect(0, A4[1]-1.2*cm, A4[0], 1.2*cm, stroke=0, fill=1)
        c.setFillColor(colors.white)
        c.setFont(FONT_BOLD, 12)
        c.drawString(
            2*cm,
            A4[1]-0.85*cm,
            localized_text("站点勘察报告", locale_code, "Site Survey Report", "Laporan Survei Situs"),
        )
        # Footer line and page number
        c.setStrokeColor(primary)
        c.setLineWidth(0.6)
        c.line(2*cm, 1.5*cm, A4[0]-2*cm, 1.5*cm)
        c.setFont(FONT_MAIN, 9)
        c.setFillColor(colors.HexColor('#888888'))
        c.drawRightString(
            A4[0]-2*cm,
            1.1*cm,
            localized_text("第 {page} 页", locale_code, "Page {page}", "Halaman {page}").format(page=d.page),
        )
        c.restoreState()

    elements: list = []

    # Cover
    elements.append(Spacer(1, 4*cm))
    elements.append(Paragraph(localized_text("站点勘察报告", locale_code, "Site Survey Report", "Laporan Survei Situs"), styles["ArchivePdfTitle"]))
    elements.append(Spacer(1, 0.6*cm))
    site_name = survey.site.site_name if survey.site else ""
    site_code = survey.site.site_code if survey.site else ""
    elements.append(Paragraph(f"{site_name} ({site_code})", styles["ArchivePdfBody"]))
    elements.append(Spacer(1, 0.2*cm))
    meta = [
        f"{localized_text('勘察日期', locale_code, 'Survey Date', 'Tanggal Survei')}：{survey.survey_date.strftime('%Y-%m-%d %H:%M') if survey.survey_date else '-'}",
        f"{localized_text('勘察人', locale_code, 'Surveyor', 'Surveyor')}：{survey.surveyor_name or '-'}",
        f"{localized_text('结论', locale_code, 'Conclusion', 'Kesimpulan')}：{survey.feasibility or '-'}",
        f"{localized_text('坐标', locale_code, 'Coordinates', 'Koordinat')}：{(survey.latitude or '-')}, {(survey.longitude or '-')} ",
        f"{localized_text('地址', locale_code, 'Address', 'Alamat')}：{survey.address or '-'}",
    ]
    for m in meta:
        elements.append(Paragraph(m, styles["ArchivePdfMeta"]))
    elements.append(PageBreak())

    # Highlighted conclusion banner
    feas = (survey.feasibility or "").strip()
    feas_map = {
        'feasible': (localized_text("可行", locale_code, "Feasible", "Layak"), colors.HexColor('#67C23A')),
        'conditionally_feasible': (localized_text("有条件可行", locale_code, "Conditionally feasible", "Layak Bersyarat"), colors.HexColor('#E6A23C')),
        'infeasible': (localized_text("不可行", locale_code, "Infeasible", "Tidak Layak"), colors.HexColor('#F56C6C')),
    }
    feas_label, feas_color = feas_map.get(feas, (feas or localized_text("未填写", locale_code, "Not filled", "Belum diisi"), primary))
    badge = Table([[create_pdf_cell(f"{localized_text('结论', locale_code, 'Conclusion', 'Kesimpulan')}：{feas_label}", styles["ArchivePdfBadge"])]], colWidths=[15*cm])
    badge.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), feas_color),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.white),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    elements.append(badge)
    elements.append(Spacer(1, 0.3*cm))

    # 概览
    elements.append(Paragraph(localized_text("概览", locale_code, "Overview", "Ringkasan"), styles["ArchivePdfH2"]))
    kv = [
        (localized_text("站点名称", locale_code, "Site Name", "Nama Situs"), site_name),
        (localized_text("站点编码", locale_code, "Site Code", "Kode Situs"), site_code),
        (localized_text("勘察日期", locale_code, "Survey Date", "Tanggal Survei"), survey.survey_date.strftime('%Y-%m-%d %H:%M') if survey.survey_date else "-"),
        (localized_text("勘察人", locale_code, "Surveyor", "Surveyor"), survey.surveyor_name or "-"),
        (localized_text("勘察人电话", locale_code, "Surveyor Phone", "Telepon Surveyor"), survey.surveyor_phone or "-"),
        (localized_text("地址", locale_code, "Address", "Alamat"), survey.address or "-"),
        (localized_text("坐标", locale_code, "Coordinates", "Koordinat"), f"{survey.latitude or ''}, {survey.longitude or ''}"),
        (localized_text("GPS精度(m)", locale_code, "GPS Accuracy (m)", "Akurasi GPS (m)"), f"{survey.gps_accuracy}" if survey.gps_accuracy is not None else "-"),
        (localized_text("风险", locale_code, "Risks", "Risiko"), survey.risks or "-"),
        (localized_text("建议", locale_code, "Recommendations", "Rekomendasi"), survey.recommendations or "-"),
    ]
    def _fmt_val(v):
        try:
            # booleans and None
            if v is None:
                return '-'
            # numbers
            if isinstance(v, (int, float)):
                return ("%.6f" % v).rstrip('0').rstrip('.') if isinstance(v, float) else str(v)
            # datetime
            if isinstance(v, datetime):
                return v.strftime('%Y-%m-%d %H:%M')
            # other
            s = str(v)
            return s if s.strip() != '' else '-'
        except Exception:
            return str(v)

    def build_kv_table(pairs):
        data = [[create_pdf_cell(k, styles["ArchivePdfHeader"]), create_pdf_cell(_fmt_val(v), styles["ArchivePdfBody"])] for k, v in pairs]
        t = Table(data, colWidths=[3*cm, 12*cm])
        t.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, accent_bg]),
            ("LINEBELOW", (0,0), (-1,-1), 0.25, colors.HexColor('#eeeeee')),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        return t
    elements.append(build_kv_table(kv))

    # 详细信息分组
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph(localized_text("场地与结构", locale_code, "Site & Structure", "Tapak & Struktur"), styles["ArchivePdfH2"]))
    elements.append(build_kv_table([
        (localized_text("站点类型", locale_code, "Site Type", "Tipe Situs"), survey.site_type),
        (localized_text("塔型", locale_code, "Tower Type", "Tipe Menara"), survey.tower_type),
        (localized_text("可用挂高(m)", locale_code, "Available Height (m)", "Tinggi Tersedia (m)"), survey.available_height_m),
        (localized_text("荷载(kg)", locale_code, "Load Capacity (kg)", "Kapasitas Beban (kg)"), survey.load_capacity_kg),
    ]))

    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph(localized_text("供电与回传", locale_code, "Power & Backhaul", "Daya & Backhaul"), styles["ArchivePdfH2"]))
    elements.append(build_kv_table([
        (localized_text("市电可用", locale_code, "Power Available", "Daya Tersedia"), localized_text("是", locale_code, "Yes", "Ya") if survey.power_available else (localized_text("否", locale_code, "No", "Tidak") if survey.power_available is not None else '-')),
        (localized_text("电源距离(m)", locale_code, "Power Distance (m)", "Jarak Daya (m)"), survey.power_distance_m),
        (localized_text("容量(kW)", locale_code, "Capacity (kW)", "Kapasitas (kW)"), survey.power_capacity_kw),
        (localized_text("接地可行", locale_code, "Grounding Feasible", "Pembumian Layak"), localized_text("是", locale_code, "Yes", "Ya") if survey.earthing_feasible else (localized_text("否", locale_code, "No", "Tidak") if survey.earthing_feasible is not None else '-')),
        (localized_text("光纤可用", locale_code, "Fiber Available", "Serat Tersedia"), localized_text("是", locale_code, "Yes", "Ya") if survey.fiber_available else (localized_text("否", locale_code, "No", "Tidak") if survey.fiber_available is not None else '-')),
        (localized_text("光纤距离(m)", locale_code, "Fiber Distance (m)", "Jarak Serat (m)"), survey.fiber_distance_m),
        (localized_text("微波LoS", locale_code, "Microwave LoS", "LoS Microwave"), localized_text("是", locale_code, "Yes", "Ya") if survey.microwave_los else (localized_text("否", locale_code, "No", "Tidak") if survey.microwave_los is not None else '-')),
        (localized_text("方位角(°)", locale_code, "Azimuth (°)", "Azimut (°)"), survey.los_azimuth_deg),
        (localized_text("距离(km)", locale_code, "Distance (km)", "Jarak (km)"), survey.los_distance_km),
    ]))

    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph(localized_text("环境与进场", locale_code, "Environment & Access", "Lingkungan & Akses"), styles["ArchivePdfH2"]))
    elements.append(build_kv_table([
        (localized_text("敏感点", locale_code, "Sensitive Points", "Titik Sensitif"), survey.sensitive_points),
        (localized_text("安全/隐患", locale_code, "Safety / Risks", "Keamanan / Risiko"), survey.safety_notes),
        (localized_text("审批/物业限制", locale_code, "Approval / Property Limits", "Persetujuan / Batas Properti"), survey.permits_constraints),
        (localized_text("进场限制", locale_code, "Access Limits", "Batas Akses"), survey.entry_constraints),
    ]))

    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph(localized_text("业主信息", locale_code, "Owner Info", "Info Pemilik"), styles["ArchivePdfH2"]))
    elements.append(build_kv_table([
        (localized_text("业主姓名", locale_code, "Owner Name", "Nama Pemilik"), survey.owner_name),
        (localized_text("业主电话", locale_code, "Owner Phone", "Telepon Pemilik"), survey.owner_phone),
        (localized_text("时间窗口", locale_code, "Time Window", "Jendela Waktu"), survey.access_time_window),
    ]))

    # 附件清单（非图片）
    attachments = (
        db.query(SiteSurveyPhoto)
        .filter(SiteSurveyPhoto.survey_id == survey_id)
        .filter(or_(SiteSurveyPhoto.mime_type == None, ~SiteSurveyPhoto.mime_type.like('image/%')))  # noqa: E711
        .order_by(SiteSurveyPhoto.created_at.asc())
        .all()
    )
    if attachments:
        elements.append(Spacer(1, 0.4*cm))
        elements.append(Paragraph(localized_text("附件清单", locale_code, "Attachment List", "Daftar Lampiran"), styles["ArchivePdfH2"]))
        rows = [[
            create_pdf_cell(localized_text("文件名", locale_code, "File Name", "Nama Berkas"), styles["ArchivePdfHeader"]),
            create_pdf_cell(localized_text("分类", locale_code, "Category", "Kategori"), styles["ArchivePdfHeader"]),
            create_pdf_cell(localized_text("类型", locale_code, "Type", "Tipe"), styles["ArchivePdfHeader"]),
            create_pdf_cell(localized_text("大小(KB)", locale_code, "Size (KB)", "Ukuran (KB)"), styles["ArchivePdfHeader"]),
        ]]
        for a in attachments:
            size_kb = None
            try:
                if a.file_path and os.path.exists(a.file_path):
                    size_kb = int(os.path.getsize(a.file_path)/1024)
            except Exception:
                size_kb = None
            rows.append([
                create_pdf_cell(os.path.basename(a.file_path) if a.file_path else (a.original_name or a.id), styles["ArchivePdfBody"]),
                create_pdf_cell(a.category or '-', styles["ArchivePdfBody"]),
                create_pdf_cell(a.mime_type or '-', styles["ArchivePdfBody"]),
                create_pdf_cell(str(size_kb or '-'), styles["ArchivePdfBody"]),
            ])
        t = Table(rows, colWidths=[8*cm, 3*cm, 3*cm, 2*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), primary), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, accent_bg]),
            ("LINEBELOW", (0,0), (-1,-1), 0.25, colors.HexColor('#eeeeee')),
            ("LEFTPADDING", (0,0), (-1,-1), 6), ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING", (0,0), (-1,-1), 6), ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        elements.append(t)

    # Photos grid（按分类分组排版，仅包含图片类型）
    photos = (
        db.query(SiteSurveyPhoto)
        .filter(SiteSurveyPhoto.survey_id == survey_id)
        .order_by(SiteSurveyPhoto.sort_order.asc().nulls_last(), SiteSurveyPhoto.created_at.asc())
        .all()
    )
    # 仅保留图片类型，避免将附件（PDF/CAD等）混入照片区
    photos = [p for p in photos if (p.mime_type or '').startswith('image/')]

    if photos:
        elements.append(Spacer(1, 0.6*cm))
        elements.append(Paragraph(localized_text("照片索引", locale_code, "Photo Index", "Indeks Foto"), styles["ArchivePdfH2"]))

        # 映射分类显示名称（与前端一致）
        def cat_label(k: str) -> str:
            mapping = {
                'overview': localized_text('全景', locale_code, 'Overview', 'Ringkasan'),
                'power': localized_text('电力/配电', locale_code, 'Power / Distribution', 'Daya / Distribusi'),
                'room': localized_text('机房/机柜', locale_code, 'Room / Cabinet', 'Ruang / Kabinet'),
                'duct': localized_text('管道/弱电', locale_code, 'Duct / Low Voltage', 'Duct / Tegangan Rendah'),
                'roof': localized_text('屋面/塔体', locale_code, 'Roof / Tower', 'Atap / Menara'),
                'hazard': localized_text('隐患', locale_code, 'Hazard', 'Bahaya'),
                'custom': localized_text('其他', locale_code, 'Other', 'Lainnya'),
                'uncategorized': localized_text('未分类', locale_code, 'Uncategorized', 'Belum dikategorikan'),
            }
            return mapping.get(k, k or localized_text('未分类', locale_code, 'Uncategorized', 'Belum dikategorikan'))

        # 分组
        grouped: dict[str, list] = {}
        for p in photos:
            k = p.category or 'uncategorized'
            grouped.setdefault(k, []).append(p)

        # 分类顺序：已知类别优先，然后其余按名称排序
        known_order = ['overview', 'power', 'room', 'duct', 'roof', 'hazard', 'custom', 'uncategorized']
        present = list(grouped.keys())
        ordered_cats = [k for k in known_order if k in present] + sorted([k for k in present if k not in known_order])

        # 逐分类排版（两列网格）
        for cat in ordered_cats:
            group = [p for p in grouped[cat] if p.file_path and os.path.exists(p.file_path)]
            if not group:
                continue
            elements.append(Paragraph(f"{cat_label(cat)}（{len(group)}）", styles["ArchivePdfH3"]))

            row: list = []
            grid: list = []
            for p in group:
                try:
                    img = Image(p.file_path, width=7.4*cm, height=5.0*cm)
                    # 标注文件名作为说明
                    caption = Paragraph(os.path.basename(p.file_path), styles["ArchivePdfCaption"])
                    cell = Table([[img], [caption]], colWidths=[7.4*cm])
                    cell.setStyle(TableStyle([
                        ("ALIGN", (0,0), (-1,-1), "CENTER"),
                        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ]))
                    row.append(cell)
                    if len(row) == 2:
                        grid.append(row)
                        row = []
                except Exception:
                    continue
            if row:
                # 填充最后一列占位，保持两列对齐
                row.append(Spacer(7.4*cm, 0))
                grid.append(row)
            if grid:
                grid_table = Table(grid, colWidths=[7.6*cm, 7.6*cm], hAlign='LEFT', spaceBefore=6)
                grid_table.setStyle(TableStyle([
                    ("VALIGN", (0,0), (-1,-1), "TOP"),
                    ("LEFTPADDING", (0,0), (-1,-1), 4),
                    ("RIGHTPADDING", (0,0), (-1,-1), 4),
                    ("TOPPADDING", (0,0), (-1,-1), 6),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ]))
                elements.append(grid_table)

    # Build
    doc.build(elements, onFirstPage=draw_page_frame, onLaterPages=draw_page_frame)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=site_survey_{survey_id}.pdf"})
