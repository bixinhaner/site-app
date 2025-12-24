from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime
from functools import lru_cache
from pathlib import Path
import math
import io
import json
import re
import pandas as pd
from openpyxl import load_workbook

from app.core.database import get_db
from app.api.auth import get_current_user
from app.models.user import User
from app.models.site import Site
from app.models.work_order import WorkOrder, WorkOrderTypeEnum, WorkOrderStatusEnum
from app.models.planning import (
    SitePlanning,
    SitePlanningSector,
    SiteAntennaPort,
    SiteSwitchPort,
    PlanningChangeLog,
    SitePlanningCell,
)
from app.utils.planning_schema import LLD_CELL_EXTRA_FIELD_CANDIDATES
from app.schemas.planning import (
    SitePlanningResponse,
    SitePlanningUpdate,
    SitePlanningVersion,
    PlanningImportReport,
    SitePlanningBase,
    PlanningChangeLogItem,
    BatchImportReport,
    BatchPlanningResult,
    PlanningCell,
    SitePlanningLldResponse,
    SitePlanningLldSummary,
    LldEditPolicy,
    PlanningCellUpdate,
    PlanningCellCreate,
    LldPlanningSummaryItem,
    LldPlanningSummaryListResponse,
    LldPlanningCellItem,
    LldPlanningCellListResponse,
    LldTemplateCellListResponse,
)


router = APIRouter()

LLD_EDIT_ALLOWED_ROLES = {"admin", "manager", "planner"}
LLD_LOCKED_FIELDS = [
    "rat",
    "band_code",
    "local_cell_id",
    "tower_id",
    "enb_id",
    "gnb_id",
    "eci",
    "nci",
    "pci",
    "frequency",
    "bandwidth",
]
LLD_FULL_EDIT_SITE_STATUSES = {"planning", "planned"}
LLD_ACTIVE_OPENING_WO_STATUSES = [
    WorkOrderStatusEnum.PENDING,
    WorkOrderStatusEnum.ACTIVE,
    WorkOrderStatusEnum.SUBMITTED,
    WorkOrderStatusEnum.UNDER_REVIEW,
    WorkOrderStatusEnum.APPROVED,
    WorkOrderStatusEnum.ACTIVATED,
]

LLD_TEMPLATE_HEADER_TO_FIELD: Dict[str, str] = {
    # base fields
    "Tower Merchants": "tower_merchants",
    "TOWER NAME": "tower_name",
    "TOWER ID": "tower_id",
    "Town": "town",
    "Sector ID Local ID": "local_cell_id",
    "Sector ID": "local_cell_id",
    "CELL NAME": "cell_name",
    "PLMN": "plmn",
    "TAC": "tac",
    "EARFCN": "frequency",
    "Bandwidth": "bandwidth",
    "Power": "power_dbm",
    "Power（dbm）": "power_dbm",
    "PA": "pa",
    "PB": "pb",
    "PCI": "pci",
    "ZC Root Index": "zc_root_index",
    "Longitude": "longitude",
    "Latitude": "latitude",
    "Tower Height": "tower_height",
    "Antenna Height": "antenna_height",
    "M-Tilt": "mechanical_downtilt_deg",
    "E-Tilt": "electrical_downtilt_deg",
    "Azimuth": "azimuth_deg",
    "Band Combination": "band_combination",
    "Cell Allocation": "cell_allocation",
    "Coverage type": "cover_type",
    "Coverage scenarios": "scenario",
    "Area priority value": "coverage_weight",
    "scenarios priority value": "scenario_weight",
    "Priority Value": "weight",
    "priority value": "weight",
    # LTE specific
    "ENB ID": "enb_id",
    "ECI": "eci",
    "MME IP 1": "mme_ip_1",
    "MME IP 2": "mme_ip_2",
    "MME IP 3": "mme_ip_3",
    "MME IP 4": "mme_ip_4",
    "MME IP 5": "mme_ip_5",
    "MME IP 6": "mme_ip_6",
    "MME IP 7": "mme_ip_7",
    "MME IP 8": "mme_ip_8",
    # NR specific
    "gNB ID": "gnb_id",
    "Gnb length": "gnb_length",
    "NCI": "nci",
    "Kssb": "kssb",
    "Offset to PointA": "offset_to_point_a",
    "Slot config": "slot_config",
    "Slot config DL/UL": "slot_config_dl_ul",
    "Symbol config DL/UL": "symbol_config_dl_ul",
    "DL SubCarrierSpacing": "dl_subcarrier_spacing",
    "MASTER 5GC IP1": "master_5gc_ip1",
    "MASTER 5GC IP2": "master_5gc_ip2",
    "MASTER 5GC IP3": "master_5gc_ip3",
    "BACKUP 5GC IP1": "backup_5gc_ip1",
    "BACKUP 5GC IP2": "backup_5gc_ip2",
    "BACKUP 5GC IP3": "backup_5gc_ip3",
    # Common IPs
    "MASTER OMC IP": "master_omc_ip",
    "BACKUP OMC IP": "backup_omc_ip",
    "NTP IP1": "ntp_ip1",
    "NTP IP2": "ntp_ip2",
    "Remark": "remark",
}


def _has_active_opening_work_order(db: Session, site_id: int) -> bool:
    if not site_id:
        return False
    exists = (
        db.query(WorkOrder.id)
        .filter(
            WorkOrder.site_id == site_id,
            WorkOrder.type == WorkOrderTypeEnum.OPENING_INSPECTION,
            WorkOrder.status.in_(LLD_ACTIVE_OPENING_WO_STATUSES),
        )
        .first()
    )
    return bool(exists)


def _active_opening_work_orders_site_ids(db: Session, site_ids: List[int]) -> set:
    ids = [int(i) for i in (site_ids or []) if i]
    if not ids:
        return set()
    rows = (
        db.query(WorkOrder.site_id)
        .filter(
            WorkOrder.site_id.in_(ids),
            WorkOrder.type == WorkOrderTypeEnum.OPENING_INSPECTION,
            WorkOrder.status.in_(LLD_ACTIVE_OPENING_WO_STATUSES),
        )
        .distinct()
        .all()
    )
    return {int(r[0]) for r in rows if r and r[0] is not None}


def _compute_lld_edit_policy(site: Site, has_active_opening_wo: bool, current_user: User) -> LldEditPolicy:
    status = (getattr(site, "status", None) or "").strip()
    has_role = bool(current_user and getattr(current_user, "role", None) in LLD_EDIT_ALLOWED_ROLES)

    if not has_role:
        return LldEditPolicy(
            mode="readonly",
            can_edit=False,
            can_import=False,
            can_add_cell=False,
            can_delete_cell=False,
            locked_fields=LLD_LOCKED_FIELDS,
            reason="无权限编辑规划",
        )

    if status == "survey_pending":
        return LldEditPolicy(
            mode="readonly",
            can_edit=False,
            can_import=False,
            can_add_cell=False,
            can_delete_cell=False,
            locked_fields=LLD_LOCKED_FIELDS,
            reason="站点尚处于勘察阶段（survey_pending），暂不允许录入/编辑规划信息",
        )

    # 规划中/规划完成：若不存在进行中的开站工单，则可全量编辑；否则降级为受限编辑
    if status in LLD_FULL_EDIT_SITE_STATUSES and not has_active_opening_wo:
        return LldEditPolicy(
            mode="full",
            can_edit=True,
            can_import=True,
            can_add_cell=True,
            can_delete_cell=True,
            locked_fields=LLD_LOCKED_FIELDS,
            reason=None,
        )

    # 其他阶段（或存在进行中开站工单）统一为受限编辑
    reason = None
    if has_active_opening_wo:
        reason = "存在进行中的开站工单，规划仅允许修改非关键信息"
    else:
        reason = f"站点当前阶段为 {status or '-'}，规划仅允许修改非关键信息"

    return LldEditPolicy(
        mode="limited",
        can_edit=True,
        can_import=True,
        can_add_cell=False,
        can_delete_cell=False,
        locked_fields=LLD_LOCKED_FIELDS,
        reason=reason,
    )


def _format_lld_cell_key(key: Dict[str, Any]) -> str:
    rat = key.get("rat") or "-"
    band = key.get("band_code") or "-"
    lcid = key.get("local_cell_id")
    lcid_str = str(lcid) if lcid is not None else "-"
    return f"{rat}/{band}/LCID={lcid_str}"


def _find_lld_policy_conflicts_for_import(policy: LldEditPolicy, diff: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """根据编辑策略检查导入带来的变更是否包含被禁止的操作。

    limited 模式下禁止：
    - Cell 增删（created/deleted）
    - 锁字段变更（LLD_LOCKED_FIELDS）
    """
    if not policy or policy.mode != "limited":
        return None

    cell_changes = (diff or {}).get("cell_changes") or []
    created = []
    deleted = []
    locked_field_changes = []

    for chg in cell_changes:
        change_type = chg.get("change_type")
        key = chg.get("key") or {}
        if change_type == "created":
            created.append({"key": key})
            continue
        if change_type == "deleted":
            deleted.append({"key": key})
            continue
        if change_type == "updated":
            for f in chg.get("changes") or []:
                field = f.get("field")
                if field in LLD_LOCKED_FIELDS:
                    locked_field_changes.append(
                        {
                            "key": key,
                            "field": field,
                            "old": f.get("old"),
                            "new": f.get("new"),
                        }
                    )

    if not created and not deleted and not locked_field_changes:
        return None

    return {
        "created": created,
        "deleted": deleted,
        "locked_field_changes": locked_field_changes,
    }


def _build_lld_policy_conflict_errors(policy: LldEditPolicy, conflicts: Dict[str, Any]) -> List[str]:
    created = conflicts.get("created") or []
    deleted = conflicts.get("deleted") or []
    locked = conflicts.get("locked_field_changes") or []

    locked_fields_text = "，".join(LLD_LOCKED_FIELDS)

    lines = [
        (policy.reason or "站点处于受限编辑状态：仅允许修改非关键信息。"),
        f"受限编辑禁止：新增/删除 Cell，修改关键字段（{locked_fields_text}）。",
    ]

    if created:
        sample = "；".join(_format_lld_cell_key(i.get("key") or {}) for i in created[:5])
        lines.append(f"检测到将新增 Cell：{len(created)} 个（示例：{sample}）")
    if deleted:
        sample = "；".join(_format_lld_cell_key(i.get("key") or {}) for i in deleted[:5])
        lines.append(f"检测到将删除 Cell：{len(deleted)} 个（示例：{sample}）")
    if locked:
        sample_items = []
        for it in locked[:5]:
            sample_items.append(
                f"{_format_lld_cell_key(it.get('key') or {})} 字段 {it.get('field')}: {it.get('old')} -> {it.get('new')}"
            )
        lines.append(f"检测到关键字段变更：{len(locked)} 处（示例：{'；'.join(sample_items)}）")

    lines.append("请修正 LLD Excel 后重试导入。")
    return lines


def _merge_lld_cell_diff_into_planning_log(db: Session, planning_id: int, cell_diff: Optional[Dict[str, Any]]) -> None:
    """将 LLD Cell 级 diff 合并写入规划变更日志（PlanningChangeLog.diff）。"""
    if not planning_id or not cell_diff:
        return
    log_entry = (
        db.query(PlanningChangeLog)
        .filter(PlanningChangeLog.planning_id == planning_id)
        .order_by(PlanningChangeLog.created_at.desc())
        .first()
    )
    if not log_entry:
        return

    merged = log_entry.diff or {}

    existing_cf = set(merged.get("changed_fields") or [])
    for f in cell_diff.get("changed_fields") or []:
        existing_cf.add(f)
    if existing_cf:
        merged["changed_fields"] = sorted(existing_cf)

    existing_cc = merged.get("cell_changes") or []
    existing_cc.extend(cell_diff.get("cell_changes") or [])
    if existing_cc:
        merged["cell_changes"] = existing_cc

    log_entry.diff = merged
    db.add(log_entry)
    db.commit()

# 反向映射：模板列名 -> cell 字段名
for _field_name, _candidates in LLD_CELL_EXTRA_FIELD_CANDIDATES.items():
    for _col in _candidates:
        if _col:
            LLD_TEMPLATE_HEADER_TO_FIELD.setdefault(str(_col), _field_name)


def _parse_band_list(band: Optional[str]) -> List[str]:
    if not band:
        return []
    items = [b.strip() for b in str(band).split(",") if str(b).strip()]
    # 去重但保持顺序
    seen = set()
    result = []
    for b in items:
        if b in seen:
            continue
        seen.add(b)
        result.append(b)
    return result


@lru_cache(maxsize=1)
def _get_lld_template_path() -> Path:
    base_dir = Path(__file__).resolve().parents[3]
    return base_dir / "LLD templateV1.0.xlsx"


@lru_cache(maxsize=1)
def _get_lld_template_headers() -> Dict[str, List[str]]:
    xlsx_path = _get_lld_template_path()
    if not xlsx_path.exists():
        raise HTTPException(
            status_code=500,
            detail="LLD 模板文件缺失：请在系统部署目录中放置 LLD templateV1.0.xlsx",
        )
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    headers: Dict[str, List[str]] = {}
    for sheet in ("4G", "5G"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            headers[sheet] = []
            continue
        headers[sheet] = [str(v).strip() if v is not None else "" for v in list(first_row)]
    return headers


def _is_blank(v: object) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _get_obj_attr(obj: object, attr: str):
    if not obj or not attr:
        return None
    v = getattr(obj, attr, None)
    if _is_blank(v):
        return None
    return v


def _to_excel_value(v: object):
    if v is None:
        return None
    if isinstance(v, (dict, list)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return v


def _get_lld_template_value(col: str, cell: SitePlanningCell, site: Site):
    # 站点字段回填：Cell 优先，空值回退 Site
    if col == "SiteCode":
        return _get_obj_attr(cell, "site_information") or _get_obj_attr(site, "site_code")
    if col == "SITE NAME":
        return _get_obj_attr(cell, "site_name") or _get_obj_attr(site, "site_name")
    if col == "Province Region":
        return (
            _get_obj_attr(cell, "province_region")
            or _get_obj_attr(cell, "region")
            or _get_obj_attr(site, "province")
        )
    if col == "Province":
        return (
            _get_obj_attr(cell, "province")
            or _get_obj_attr(cell, "region")
            or _get_obj_attr(site, "province")
        )
    if col == "City":
        return _get_obj_attr(cell, "city") or _get_obj_attr(site, "city")
    if col == "County":
        return _get_obj_attr(cell, "county") or _get_obj_attr(site, "district")
    if col == "Address":
        return _get_obj_attr(cell, "address") or _get_obj_attr(site, "address")

    # 关键字段：Band / DL带宽回退
    if col in ("BAND", "Band"):
        return _get_obj_attr(cell, "band_in_file") or _get_obj_attr(cell, "band_code")
    if col == "DL Bandwidth":
        return _get_obj_attr(cell, "dl_bandwidth") or _get_obj_attr(cell, "bandwidth")

    field = LLD_TEMPLATE_HEADER_TO_FIELD.get(col)
    if field:
        return _get_obj_attr(cell, field)

    # 未知字段：尝试从 extra_params 中取
    extra = getattr(cell, "extra_params", None)
    if isinstance(extra, dict) and col in extra and not _is_blank(extra.get(col)):
        return extra.get(col)

    return None


def _get_current_planning(db: Session, site_id: int) -> Optional[SitePlanning]:
    return (
        db.query(SitePlanning)
        .filter(SitePlanning.site_id == site_id)
        .filter(SitePlanning.is_current == True)
        .first()
    )


def _max_version(db: Session, site_id: int) -> int:
    version = db.query(func.max(SitePlanning.version)).filter(SitePlanning.site_id == site_id).scalar()
    return int(version or 0)


def _snapshot(planning: Optional[SitePlanning]) -> Optional[dict]:
    if not planning:
        return None
    return {
        "planning": {
            "bands": planning.bands or [],
            "sector_count": planning.sector_count or 0,
            "notes": planning.notes or "",
        },
        "sectors": [
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in (planning.sectors or [])
        ],
        "antenna_ports": [
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in (planning.antenna_ports or [])
        ],
        "switch_ports": [
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in (planning.switch_ports or [])
        ],
    }


def _cell_kwargs_from_model(cell: SitePlanningCell) -> dict:
    """将 SitePlanningCell ORM 对象转换为可用于新建行的 kwargs（自动涵盖新增列）。"""
    exclude = {"id", "planning_id", "created_at"}
    return {
        c.name: getattr(cell, c.name)
        for c in SitePlanningCell.__table__.columns
        if c.name not in exclude
    }


def _compute_diff(before: Optional[dict], after: Optional[dict]) -> dict:
    """
    生成规划变更的简要 diff 信息。

    当前前端只使用 diff.changed_fields 在“日志”Tab 中展示发生变更的模块，
    因此这里主要关注哪些大类发生了变化：planning / sectors / antenna_ports / switch_ports。
    """
    keys = ["planning", "sectors", "antenna_ports", "switch_ports"]

    if before is None and after is None:
        return {}

    # 首次创建或被删除时，也补充 changed_fields，便于前端展示
    if before is None:
        return {"action": "created", "changed_fields": keys}
    if after is None:
        return {"action": "deleted", "changed_fields": keys}

    diff: Dict[str, Any] = {"changed_fields": []}
    for key in keys:
        if (before or {}).get(key) != (after or {}).get(key):
            diff["changed_fields"].append(key)
    return diff


def _compute_lld_cells_diff(
    old_cells: List[SitePlanningCell],
    new_cells: List[dict],
) -> Dict[str, Any]:
    """
    针对 LLD 导入场景，比较导入前后的 Cell 明细差异。

    返回结构示例：
    {
      "changed_fields": ["cells", "cells.tac", "cells.pci"],
      "cell_changes": [
         {
           "key": {"rat": "LTE", "band_code": "B3", "local_cell_id": 1},
           "change_type": "updated",
           "changes": [
              {"field": "tac", "old": "12345", "new": "23456"},
              ...
           ]
         },
         ...
      ]
    }
    """

    def make_key(obj) -> tuple:
        rat = getattr(obj, "rat", None) if isinstance(obj, SitePlanningCell) else obj.get("rat")
        band = getattr(obj, "band_code", None) if isinstance(obj, SitePlanningCell) else obj.get("band_code")
        lcid = getattr(obj, "local_cell_id", None) if isinstance(obj, SitePlanningCell) else obj.get("local_cell_id")
        return (
            str(rat) if rat is not None else None,
            str(band) if band is not None else None,
            int(lcid) if lcid is not None else None,
        )

    # 需要比较的字段：排除 id/planning_id/site_id/timestamps 等
    exclude = {"id", "planning_id", "site_id", "created_at"}
    compare_fields = [
        c.name
        for c in SitePlanningCell.__table__.columns
        if c.name not in exclude
    ]

    old_map: Dict[tuple, SitePlanningCell] = {}
    for c in old_cells:
        k = make_key(c)
        old_map[k] = c

    new_map: Dict[tuple, dict] = {}
    for d in new_cells:
        k = make_key(d)
        if k[0] is None or k[2] is None:
            # 缺少关键信息的行（如没有 RAT 或 Sector ID）跳过
            continue
        new_map[k] = d

    all_keys = set(old_map.keys()) | set(new_map.keys())
    changed_fields: set = set()
    cell_changes: List[Dict[str, Any]] = []

    for key in all_keys:
        old = old_map.get(key)
        new = new_map.get(key)

        if old is None and new is not None:
            # 新增的 Cell
            changes = []
            for field in compare_fields:
                new_val = new.get(field)
                if new_val not in (None, ""):
                    changes.append({"field": field, "old": None, "new": new_val})
                    changed_fields.add(f"cells.{field}")
            if changes:
                cell_changes.append(
                    {
                        "key": {
                            "rat": key[0],
                            "band_code": key[1],
                            "local_cell_id": key[2],
                        },
                        "change_type": "created",
                        "changes": changes,
                    }
                )
            continue

        if old is not None and new is None:
            # 删除的 Cell
            changes = []
            for field in compare_fields:
                old_val = getattr(old, field, None)
                if old_val not in (None, ""):
                    changes.append({"field": field, "old": old_val, "new": None})
                    changed_fields.add(f"cells.{field}")
            if changes:
                cell_changes.append(
                    {
                        "key": {
                            "rat": key[0],
                            "band_code": key[1],
                            "local_cell_id": key[2],
                        },
                        "change_type": "deleted",
                        "changes": changes,
                    }
                )
            continue

        # 更新的 Cell：逐字段对比
        per_cell_changes = []
        for field in compare_fields:
            old_val = getattr(old, field, None)
            new_val = new.get(field)
            if old_val != new_val:
                per_cell_changes.append({"field": field, "old": old_val, "new": new_val})
                changed_fields.add(f"cells.{field}")

        if per_cell_changes:
            cell_changes.append(
                {
                    "key": {
                        "rat": key[0],
                        "band_code": key[1],
                        "local_cell_id": key[2],
                    },
                    "change_type": "updated",
                    "changes": per_cell_changes,
                }
            )

    result: Dict[str, Any] = {}
    if changed_fields:
        # 始终包含顶层 cells 标记，方便前端快速识别“有小区变更”
        changed_fields.add("cells")
        result["changed_fields"] = sorted(changed_fields)
    if cell_changes:
        result["cell_changes"] = cell_changes
    return result


def _ensure_site_plannable(site: Site):
    """业务门禁：仅允许在勘察完成后的状态进行规划。

    规则：当站点状态为 survey_pending（勘察前/进行中）时，禁止录入/导入规划。
    允许状态示例：planning, construction, operational, maintenance 等。
    """
    status = getattr(site, 'status', None)
    if status == 'survey_pending':
        raise HTTPException(status_code=409, detail="站点尚处于勘察阶段（survey_pending），暂不允许录入规划信息。请完成勘察后再试。")


def _normalize_tower_id(value) -> Optional[str]:
    """规范化 Excel 中的 TOWER ID（塔 ID）。"""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _normalize_site_code(value) -> Optional[str]:
    """将 Excel 中的 SiteCode 规范化为用于匹配站点的站点编码（sites.site_code）。"""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _normalize_lld_column_name(col: object) -> str:
    """规范化 LLD 模板列名。

    新版模板会在列名中用换行标注“新增/改/填写值变化”等说明，这里将其去除，
    并把多行标题合并为一行，方便后续字段匹配。
    """
    raw = "" if col is None else str(col)
    raw = raw.replace("\r", "\n").strip()
    if not raw:
        return ""

    # 只移除纯“标注行”，保留真实字段名（如 Province/Region 这种改名提示会保留两段）
    annotation_tokens = {
        "新增",
        "改",
        "修改",
        "修改名字",
        "改名",
        "同",
        "同现存",
        "填写值变化",
    }
    parts = [p.strip() for p in raw.split("\n") if p and str(p).strip()]
    cleaned = [p for p in parts if p not in annotation_tokens]
    name = " ".join(cleaned) if cleaned else raw
    # 统一多空格
    name = " ".join(name.split())
    return name


def _first_not_null(row: dict, candidates) -> Optional[object]:
    """从多列候选中取第一个非空值（排除 NaN 和空字符串）。"""
    for name in candidates:
        if name in row:
            v = row.get(name)
            if pd.isna(v):
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return v
    return None


def _to_str(v) -> Optional[str]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s if s else None


def _to_clean_str(v) -> Optional[str]:
    """将 Excel 单元格值转换为“干净”的字符串（用于模板扩展字段：统一按字符串保存）。"""
    sv = _sanitize_excel_value(v)
    if sv is None:
        return None
    if isinstance(sv, bool):
        return "true" if sv else "false"
    if isinstance(sv, int):
        return str(sv)
    if isinstance(sv, float):
        if sv.is_integer():
            return str(int(sv))
        return str(sv)
    if isinstance(sv, (dict, list)):
        try:
            import json

            return json.dumps(sv, ensure_ascii=False)
        except Exception:
            return str(sv)
    s = str(sv).strip()
    return s if s else None


def _to_int(v) -> Optional[int]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(str(v).split('.')[0])
    except Exception:
        return None


def _to_float(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return float(v)
    except Exception:
        return None


def _sanitize_excel_value(v):
    """将 Excel 单元格值转换为可 JSON 存储的基础类型。"""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    # pandas Timestamp / datetime
    try:
        import pandas as _pd
        if isinstance(v, _pd.Timestamp):
            return v.to_pydatetime().isoformat()
    except Exception:
        pass
    if isinstance(v, datetime):
        return v.isoformat()
    # numpy scalar
    if hasattr(v, "item") and callable(getattr(v, "item")):
        try:
            return v.item()
        except Exception:
            pass
    return v


def _create_new_version(
    db: Session,
    site_id: int,
    data: SitePlanningBase,
    actor_id: int,
    operation: str,
    summary: Optional[str] = None,
) -> SitePlanning:
    current = _get_current_planning(db, site_id)
    before_snapshot = _snapshot(current)

    # mark previous current to False
    if current:
        current.is_current = False
        db.flush()

    # 确保站点状态从 planning 切换为 planned（仅在第一次形成规划基线时）
    site = db.query(Site).filter(Site.id == site_id).first()
    if site and site.status == "planning":
        site.status = "planned"

    new_version = _max_version(db, site_id) + 1
    planning = SitePlanning(
        site_id=site_id,
        version=new_version,
        bands=data.bands,
        sector_count=data.sector_count,
        notes=data.notes,
        is_current=True,
        created_by=actor_id,
    )
    db.add(planning)
    db.flush()

    # child rows
    for s in data.sectors:
        db.add(
            SitePlanningSector(
                planning_id=planning.id,
                sector_index=s.sector_index,
                azimuth_deg=s.azimuth_deg,
                downtilt_deg=s.downtilt_deg,
                bands=s.bands,
            )
        )
    for p in data.antenna_ports:
        db.add(
            SiteAntennaPort(
                planning_id=planning.id,
                port_label=p.port_label,
                sector_index=p.sector_index,
                band=p.band,
                mimo_chain=p.mimo_chain,
                remarks=p.remarks,
            )
        )
    for sp in data.switch_ports:
        db.add(
            SiteSwitchPort(
                planning_id=planning.id,
                port_no=sp.port_no,
                vlan_ids=sp.vlan_ids,
                is_uplink=sp.is_uplink,
                poe=sp.poe,
                description=sp.description,
            )
        )

    db.commit()
    db.refresh(planning)

    after_snapshot = _snapshot(planning)
    diff = _compute_diff(before_snapshot, after_snapshot)

    # 对 LLD 相关操作，标记 cells 发生变化，便于前端在“日志”Tab 中区分
    if operation.startswith("lld"):
        changed = diff.setdefault("changed_fields", [])
        if "cells" not in changed:
            changed.append("cells")

    log = PlanningChangeLog(
        site_id=site_id,
        planning_id=planning.id,
        operation=operation,
        actor_id=actor_id,
        summary=summary,
        before_snapshot=before_snapshot,
        after_snapshot=after_snapshot,
        diff=diff,
    )
    db.add(log)
    db.commit()

    return planning


def _collect_lld_rows_by_tower(excel: pd.ExcelFile) -> dict:
    """
    按 SiteCode 聚合 LLD 行数据（用于匹配 sites.site_code）。

    返回结构:
        {
          "SiteCode": [
              {"rat": "LTE"|"NR", "band_code": "B20"/"N41", "sheet_name": "4G"/"5G", "row": {...}},
              ...
          ],
          ...
        }
    """
    rows_by_tower: Dict[str, List[dict]] = {}

    # 仅支持新模板：
    # - Sheet：4G/5G
    # - 频段从行内 BAND/Band 列读取
    #
    # 若检测到旧模板的 Sheet（如 B28/N41），直接报错提示使用新模板。
    legacy_sheets: List[str] = []
    for sheet in excel.sheet_names:
        if not sheet:
            continue
        s = str(sheet).strip()
        if not s:
            continue
        if re.match(r"^[BN]\d+$", s.upper()):
            legacy_sheets.append(s)
    if legacy_sheets:
        raise HTTPException(
            status_code=400,
            detail=f"检测到旧版 LLD 模板 Sheet：{', '.join(sorted(set(legacy_sheets)))}。旧模板已停用，请使用新模板（Sheet：4G/5G）重新导入。",
        )

    for sheet_name in excel.sheet_names:
        if not sheet_name:
            continue
        sname = str(sheet_name).strip()
        if not sname:
            continue
        upper = sname.upper()

        df = excel.parse(sheet_name)
        if df.empty:
            continue
        df.rename(columns=_normalize_lld_column_name, inplace=True)

        def add_row(row_dict: dict, rat: str, band_code: Optional[str], sheet_label: str):
            site_code = _normalize_site_code(_first_not_null(row_dict, ["SiteCode"]))
            local_cell_val = _first_not_null(row_dict, ["Sector ID Local ID", "Sector ID"])
            if not site_code or local_cell_val is None:
                return
            tower_id = _normalize_tower_id(_first_not_null(row_dict, ["TOWER ID"]))
            b = _to_str(band_code)
            b = b.upper() if b else None
            rows_by_tower.setdefault(site_code, []).append(
                {
                    "rat": rat,
                    "band_code": b or sheet_label,
                    "sheet_name": sheet_label,
                    "tower_id": tower_id,
                    "row": row_dict,
                }
            )

        if upper in ("4G", "4G CELL", "LTE"):
            if "SiteCode" not in df.columns:
                raise HTTPException(
                    status_code=400,
                    detail="LLD 模板缺少 SiteCode 列（原 SITE INFORMATION 已改名为 SiteCode），请下载最新模板重新导入。",
                )
            for _, r in df.iterrows():
                row_dict = r.to_dict()
                band = _to_str(_first_not_null(row_dict, ["BAND", "Band"]))
                # 新模板 4G/5G Sheet 不含频段信息的情况下无法推断 band_code，直接跳过
                if not band:
                    continue
                add_row(row_dict, rat="LTE", band_code=band, sheet_label=sname)
            continue

        if upper in ("5G", "5G CELL", "NR"):
            if "SiteCode" not in df.columns:
                raise HTTPException(
                    status_code=400,
                    detail="LLD 模板缺少 SiteCode 列（原 SITE INFORMATION 已改名为 SiteCode），请下载最新模板重新导入。",
                )
            for _, r in df.iterrows():
                row_dict = r.to_dict()
                band = _to_str(_first_not_null(row_dict, ["Band", "BAND"]))
                if not band:
                    continue
                add_row(row_dict, rat="NR", band_code=band, sheet_label=sname)
            continue

        # 其他 sheet 忽略
        continue

    return rows_by_tower


def _build_cell_dict_from_row(
    site_id: int,
    tower_id: Optional[str],
    meta: dict,
) -> dict:
    """将 LLD 行解析为 SitePlanningCell 可用的字段字典（不含 planning_id）。"""
    rat = meta["rat"]
    band_code = meta["band_code"]
    sheet_name = meta["sheet_name"]
    row = meta["row"]

    # 新模板列名
    local_cell_candidates = ["Sector ID Local ID", "Sector ID"]
    band_candidates = ["BAND", "Band"]

    known_cols = set(
        [
            "TOWER ID",
            "SiteCode",
            "SITE NAME",
            "CELL NAME",
            "ENB ID",
            "ECI",
            "PLMN",
            "TAC",
            "PCI",
            "ZC Root Index",
            "Longitude",
            "Latitude",
            "Power（dbm）",
            "Power",
            "PA",
            "PB",
            "Coverage type",
            "EARFCN",
            "Bandwidth",
            "DL Bandwidth",
            "M-Tilt",
            "E-Tilt",
            "Azimuth",
            "Tower Height",
            "Antenna Height",
            "Tower Merchants",
            "Band Combination",
            "Antenna port",
            "Cell Allocation",
            "TOWER NAME",
            "Town",
            "Province Region",
            "Province",
            "Coverage scenarios",
            "Area priority value",
            "scenarios priority value",
            "Priority Value",
            "priority value",
            "Remark",
            # 5G
            "gNB ID",
            "Gnb length",
            "NCI",
            "MASTER 5GC IP1",
            "MASTER 5GC IP2",
            "MASTER 5GC IP3",
            "BACKUP 5GC IP1",
            "BACKUP 5GC IP2",
            "BACKUP 5GC IP3",
            "MASTER OMC IP",
            "BACKUP OMC IP",
            "NTP IP1",
            "NTP IP2",
            "Kssb",
            "Offset to PointA",
            "Slot config",
            "Slot config DL/UL",
            "Symbol config DL/UL",
        ]
        + local_cell_candidates
        + band_candidates
    )
    # 新版模板扩展字段：不再放入 extra_params
    for candidates in LLD_CELL_EXTRA_FIELD_CANDIDATES.values():
        known_cols.update(candidates)

    extra_params = {}
    for k, v in (row or {}).items():
        if k in known_cols:
            continue
        sv = _sanitize_excel_value(v)
        if sv is None:
            continue
        if isinstance(sv, str) and not sv.strip():
            continue
        extra_params[k] = sv

    cell = {
        "site_id": site_id,
        "rat": rat,
        "band_code": band_code,
        "sheet_name": sheet_name,
        "tower_id": tower_id,
        "site_information": _to_str(_first_not_null(row, ["SiteCode"])),
        "site_name": _to_str(_first_not_null(row, ["SITE NAME"])),
        "local_cell_id": _to_int(_first_not_null(row, local_cell_candidates)),
        "cell_name": _to_str(_first_not_null(row, ["CELL NAME"])),
        "enb_id": _to_int(_first_not_null(row, ["ENB ID"])),
        "eci": _to_int(_first_not_null(row, ["ECI"])),
        "plmn": _to_str(_first_not_null(row, ["PLMN"])),
        "tac": _to_str(_first_not_null(row, ["TAC"])),
        "pci": _to_int(_first_not_null(row, ["PCI"])),
        "zc_root_index": _to_int(_first_not_null(row, ["ZC Root Index"])),
        "longitude": _to_float(_first_not_null(row, ["Longitude"])),
        "latitude": _to_float(_first_not_null(row, ["Latitude"])),
        "power_dbm": _to_float(_first_not_null(row, ["Power（dbm）", "Power"])),
        "pa": _to_str(_first_not_null(row, ["PA"])),
        "pb": _to_str(_first_not_null(row, ["PB"])),
        "cover_type": _to_str(_first_not_null(row, ["Coverage type"])),
        "band_in_file": _to_str(_first_not_null(row, band_candidates)),
        "frequency": _to_int(_first_not_null(row, ["EARFCN"])),
        # 对 5G 新模板，优先使用 DL Bandwidth；否则回退 Bandwidth
        "bandwidth": _to_str(_first_not_null(row, ["DL Bandwidth", "Bandwidth"])),
        "mechanical_downtilt_deg": _to_float(_first_not_null(row, ["M-Tilt"])),
        "electrical_downtilt_deg": _to_float(_first_not_null(row, ["E-Tilt"])),
        "azimuth_deg": _to_float(_first_not_null(row, ["Azimuth"])),
        "tower_height": _to_float(_first_not_null(row, ["Tower Height"])),
        "antenna_height": _to_float(_first_not_null(row, ["Antenna Height"])),
        "tower_merchants": _to_str(_first_not_null(row, ["Tower Merchants"])),
        "band_combination": _to_str(_first_not_null(row, ["Band Combination"])),
        "antenna_ports": _to_int(_first_not_null(row, ["Antenna port"])),
        "cell_allocation": _to_str(_first_not_null(row, ["Cell Allocation"])),
        "tower_name": _to_str(_first_not_null(row, ["TOWER NAME"])),
        "town": _to_str(_first_not_null(row, ["Town"])),
        "region": _to_str(_first_not_null(row, ["Province Region", "Province"])),
        "coverage_area": None,
        "coverage_weight": _to_str(_first_not_null(row, ["Area priority value"])),
        "scenario": _to_str(_first_not_null(row, ["Coverage scenarios"])),
        "scenario_weight": _to_str(_first_not_null(row, ["scenarios priority value"])),
        "weight": _to_str(_first_not_null(row, ["Priority Value", "priority value"])),
        "remark": _to_str(_first_not_null(row, ["Remark"])),
        # 5G 字段（如果列不存在或为 NaN，会自动为 None）
        "gnb_id": _to_int(_first_not_null(row, ["gNB ID"])),
        "gnb_length": _to_int(_first_not_null(row, ["Gnb length"])),
        "nci": _to_int(_first_not_null(row, ["NCI"])),
        "gnb_wan_ip": None,
        "master_5gc_ip1": _to_str(_first_not_null(row, ["MASTER 5GC IP1"])),
        "master_5gc_ip2": _to_str(_first_not_null(row, ["MASTER 5GC IP2"])),
        "master_5gc_ip3": _to_str(_first_not_null(row, ["MASTER 5GC IP3"])),
        "backup_5gc_ip1": _to_str(_first_not_null(row, ["BACKUP 5GC IP1"])),
        "backup_5gc_ip2": _to_str(_first_not_null(row, ["BACKUP 5GC IP2"])),
        "backup_5gc_ip3": _to_str(_first_not_null(row, ["BACKUP 5GC IP3"])),
        "master_omc_ip": _to_str(_first_not_null(row, ["MASTER OMC IP"])),
        "backup_omc_ip": _to_str(_first_not_null(row, ["BACKUP OMC IP"])),
        "ntp_ip1": _to_str(_first_not_null(row, ["NTP IP1"])),
        "ntp_ip2": _to_str(_first_not_null(row, ["NTP IP2"])),
        "kssb": _to_float(_first_not_null(row, ["Kssb"])),
        "offset_to_point_a": _to_str(_first_not_null(row, ["Offset to PointA"])),
        "slot_config": _to_str(_first_not_null(row, ["Slot config"])),
        "slot_config_dl_ul": _to_str(_first_not_null(row, ["Slot config DL/UL"])),
        "symbol_config_dl_ul": _to_str(_first_not_null(row, ["Symbol config DL/UL"])),
        # 新版模板扩展字段（统一字符串）
        **{
            field_name: _to_clean_str(_first_not_null(row, candidates))
            for field_name, candidates in LLD_CELL_EXTRA_FIELD_CANDIDATES.items()
        },
        "extra_params": extra_params or None,
    }
    return cell


def _build_planning_from_cells(site_id: int, cells: List[dict]) -> SitePlanningBase:
    """根据 Cell 明细聚合生成 SitePlanningBase（bands + sectors）。"""
    # bands: 使用 band_code 去重
    bands = sorted(
        {c["band_code"] for c in cells if c.get("band_code")}
    )

    # sectors: 以 Sector ID 作为扇区号
    sector_ids = sorted(
        {c["local_cell_id"] for c in cells if c.get("local_cell_id") is not None}
    )
    sectors = []
    for sid in sector_ids:
        sec_cells = [c for c in cells if c.get("local_cell_id") == sid]
        # 取第一个有方位角/下倾角的 Cell
        az = None
        mech = None
        elec = None
        for c in sec_cells:
            if az is None and c.get("azimuth_deg") is not None:
                az = c["azimuth_deg"]
            if mech is None and c.get("mechanical_downtilt_deg") is not None:
                mech = c["mechanical_downtilt_deg"]
            if elec is None and c.get("electrical_downtilt_deg") is not None:
                elec = c["electrical_downtilt_deg"]
        downtilt = (mech or 0.0) + (elec or 0.0)
        sector_bands = sorted(
            {c["band_code"] for c in sec_cells if c.get("band_code")}
        )
        sectors.append(
            {
                "sector_index": int(sid),
                "azimuth_deg": float(az or 0.0),
                "downtilt_deg": float(downtilt),
                "bands": sector_bands,
            }
        )

    sector_count = len(sector_ids)

    return SitePlanningBase(
        bands=bands,
        sector_count=sector_count,
        notes=None,
        sectors=sectors,
        antenna_ports=[],
        switch_ports=[],
    )


def _build_lld_summary_from_cells(planning: SitePlanning, cells: List[SitePlanningCell]) -> SitePlanningLldSummary:
    """根据当前规划版本与 Cell 明细构建概要统计。"""
    bands = sorted({c.band_code for c in cells if c.band_code})
    lte_cell_count = sum(1 for c in cells if c.rat == "LTE")
    nr_cell_count = sum(1 for c in cells if c.rat == "NR")

    mech_vals = [c.mechanical_downtilt_deg for c in cells if c.mechanical_downtilt_deg is not None]
    elec_vals = [c.electrical_downtilt_deg for c in cells if c.electrical_downtilt_deg is not None]

    return SitePlanningLldSummary(
        bands=bands,
        sector_count=planning.sector_count or 0,
        lte_cell_count=lte_cell_count,
        nr_cell_count=nr_cell_count,
        mechanical_downtilt_min=min(mech_vals) if mech_vals else None,
        mechanical_downtilt_max=max(mech_vals) if mech_vals else None,
        electrical_downtilt_min=min(elec_vals) if elec_vals else None,
        electrical_downtilt_max=max(elec_vals) if elec_vals else None,
    )


def _sync_planning_from_cells(db: Session, planning: SitePlanning, cells: List[SitePlanningCell]):
    """根据 Cell 明细同步更新基础规划数据，确保数据一致性。"""
    if not cells:
        return planning

    # 从 Cells 聚合基础规划数据
    bands = sorted({c.band_code for c in cells if c.band_code})

    # 获取唯一的扇区ID（基于 Sector ID）
    sector_ids = sorted({c.local_cell_id for c in cells if c.local_cell_id is not None})

    # 构建或更新扇区数据
    sectors = []
    for sector_id in sector_ids:
        sector_cells = [c for c in cells if c.local_cell_id == sector_id]

        # 从该扇区的Cell中获取方位角和下倾角
        azimuth = 0.0
        mechanical_downtilt = 0.0
        electrical_downtilt = 0.0
        sector_bands = []

        for cell in sector_cells:
            if cell.azimuth_deg is not None:
                azimuth = float(cell.azimuth_deg)
            if cell.mechanical_downtilt_deg is not None:
                mechanical_downtilt = float(cell.mechanical_downtilt_deg)
            if cell.electrical_downtilt_deg is not None:
                electrical_downtilt = float(cell.electrical_downtilt_deg)
            if cell.band_code:
                sector_bands.append(cell.band_code)

        total_downtilt = mechanical_downtilt + electrical_downtilt

        sectors.append({
            "sector_index": int(sector_id),
            "azimuth_deg": azimuth,
            "downtilt_deg": total_downtilt,
            "bands": sorted(set(sector_bands)),
        })

    # 更新基础规划数据
    planning.bands = bands
    planning.sector_count = len(sector_ids)

    # 删除旧的扇区数据并重新创建
    for sector in planning.sectors or []:
        db.delete(sector)

    for sector_data in sectors:
        db.add(SitePlanningSector(
            planning_id=planning.id,
            sector_index=sector_data["sector_index"],
            azimuth_deg=sector_data["azimuth_deg"],
            downtilt_deg=sector_data["downtilt_deg"],
            bands=sector_data["bands"],
        ))

    db.flush()
    return planning


def _validate_cell_data(cell_data: dict, operation: str = "create") -> List[str]:
    """验证Cell数据的业务规则，返回错误信息列表。"""
    errors = []

    # RAT 验证（如果是更新且RAT不在更新数据中，则跳过）
    rat = cell_data.get("rat")
    if rat is not None and rat not in ["LTE", "NR"]:
        errors.append("RAT必须是LTE或NR")

    # 频段验证（如果是更新且band_code不在更新数据中，则跳过）
    band_code = cell_data.get("band_code")
    if band_code is not None and not band_code:
        errors.append("频段代码不能为空")

    # PCI验证 (如果提供)
    pci = cell_data.get("pci")
    if pci is not None:
        if not isinstance(pci, int) or pci < 0 or pci > 503:
            errors.append("PCI必须是0-503之间的整数")

    # 经纬度验证
    longitude = cell_data.get("longitude")
    if longitude is not None and (longitude < -180 or longitude > 180):
        errors.append("经度必须在-180到180之间")

    latitude = cell_data.get("latitude")
    if latitude is not None and (latitude < -90 or latitude > 90):
        errors.append("纬度必须在-90到90之间")

    # 功率验证
    power_dbm = cell_data.get("power_dbm")
    if power_dbm is not None and (power_dbm < -50 or power_dbm > 80):
        errors.append("功率必须在-50到80dBm之间")

    # 角度验证
    azimuth_deg = cell_data.get("azimuth_deg")
    if azimuth_deg is not None and (azimuth_deg < 0 or azimuth_deg > 360):
        errors.append("方位角必须在0到360度之间")

    mechanical_downtilt_deg = cell_data.get("mechanical_downtilt_deg")
    if mechanical_downtilt_deg is not None and (mechanical_downtilt_deg < 0 or mechanical_downtilt_deg > 90):
        errors.append("机械下倾角必须在0到90度之间")

    electrical_downtilt_deg = cell_data.get("electrical_downtilt_deg")
    if electrical_downtilt_deg is not None and (electrical_downtilt_deg < 0 or electrical_downtilt_deg > 90):
        errors.append("电子下倾角必须在0到90度之间")

    # Local Cell ID验证
    local_cell_id = cell_data.get("local_cell_id")
    if local_cell_id is not None:
        if not isinstance(local_cell_id, int) or local_cell_id < 1 or local_cell_id > 65535:
            errors.append("Local Cell ID必须是1-65535之间的整数")

    return errors


def _check_cell_conflicts(db: Session, site_id: int, cell_data: dict, exclude_cell_id: Optional[int] = None) -> List[str]:
    """检查Cell冲突（相同频段和扇区的重复Cell）。"""
    conflicts = []

    local_cell_id = cell_data.get("local_cell_id")
    band_code = cell_data.get("band_code")
    rat = cell_data.get("rat")

    if local_cell_id is None or band_code is None or rat is None:
        return conflicts

    # 查找当前规划版本
    current_planning = _get_current_planning(db, site_id)
    if not current_planning:
        return conflicts

    # 查找冲突的Cell
    conflict_cells = (
        db.query(SitePlanningCell)
        .filter(
            SitePlanningCell.planning_id == current_planning.id,
            SitePlanningCell.site_id == site_id,
            SitePlanningCell.local_cell_id == local_cell_id,
            SitePlanningCell.band_code == band_code,
            SitePlanningCell.rat == rat,
        )
    )

    if exclude_cell_id:
        conflict_cells = conflict_cells.filter(SitePlanningCell.id != exclude_cell_id)

    conflict_cells = conflict_cells.all()

    if conflict_cells:
        conflicts.append(f"已存在相同的Cell: 扇区{local_cell_id} - {rat} {band_code}")

    return conflicts


@router.get("/{site_id}/planning", response_model=SitePlanningResponse)
async def get_current_planning(site_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    planning = _get_current_planning(db, site_id)
    if not planning:
        # return an empty frame with version 0
        raise HTTPException(status_code=404, detail="Planning not found for this site")
    # Eager load children by access
    _ = planning.sectors, planning.antenna_ports, planning.switch_ports
    return SitePlanningResponse(
        id=planning.id,
        site_id=site_id,
        version=planning.version,
        is_current=planning.is_current,
        bands=planning.bands or [],
        sector_count=planning.sector_count or 0,
        notes=planning.notes,
        sectors=[
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in planning.sectors
        ],
        antenna_ports=[
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in planning.antenna_ports
        ],
        switch_ports=[
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in planning.switch_ports
        ],
    )


@router.put("/{site_id}/planning", response_model=SitePlanningResponse)
async def update_planning(
    site_id: int,
    payload: SitePlanningUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Permissions: admin/manager can write; inspectors/users read only
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    # 生命周期门禁
    _ensure_site_plannable(site)

    # optimistic lock
    current = _get_current_planning(db, site_id)
    if payload.base_version is not None and current and current.version != payload.base_version:
        raise HTTPException(status_code=409, detail=f"Version conflict: current={current.version}, base={payload.base_version}")

    planning = _create_new_version(db, site_id, payload, current_user.id, operation="update", summary="Manual update")

    return await get_current_planning(site_id, db, current_user)


@router.get("/{site_id}/planning/versions", response_model=List[SitePlanningVersion])
async def list_versions(site_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    plans = (
        db.query(SitePlanning)
        .filter(SitePlanning.site_id == site_id)
        .order_by(SitePlanning.version.desc())
        .all()
    )
    return [
        SitePlanningVersion(
            id=p.id,
            site_id=site_id,
            version=p.version,
            is_current=p.is_current,
            created_by=p.created_by,
            created_at=p.created_at.isoformat() if p.created_at else None,
            notes=p.notes,
        )
        for p in plans
    ]


@router.get("/{site_id}/planning/versions/{version}", response_model=SitePlanningResponse)
async def get_version(site_id: int, version: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    planning = (
        db.query(SitePlanning)
        .filter(SitePlanning.site_id == site_id, SitePlanning.version == version)
        .first()
    )
    if not planning:
        raise HTTPException(status_code=404, detail="Version not found")
    _ = planning.sectors, planning.antenna_ports, planning.switch_ports
    return await get_current_planning(site_id, db, current_user) if planning.is_current else SitePlanningResponse(
        id=planning.id,
        site_id=site_id,
        version=planning.version,
        is_current=planning.is_current,
        bands=planning.bands or [],
        sector_count=planning.sector_count or 0,
        notes=planning.notes,
        sectors=[
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in planning.sectors
        ],
        antenna_ports=[
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in planning.antenna_ports
        ],
        switch_ports=[
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in planning.switch_ports
        ],
    )


@router.post("/{site_id}/planning/versions/{version}/restore", response_model=SitePlanningResponse)
async def restore_version(site_id: int, version: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    planning = (
        db.query(SitePlanning)
        .filter(SitePlanning.site_id == site_id, SitePlanning.version == version)
        .first()
    )
    if not planning:
        raise HTTPException(status_code=404, detail="Version not found")

    # build payload from snapshot
    snap = _snapshot(planning)
    data = SitePlanningBase(
        bands=snap["planning"]["bands"],
        sector_count=snap["planning"]["sector_count"],
        notes=snap["planning"].get("notes"),
        sectors=snap["sectors"],
        antenna_ports=snap["antenna_ports"],
        switch_ports=snap["switch_ports"],
    )
    # 创建新的规划版本
    new_planning = _create_new_version(
        db,
        site_id,
        data,
        current_user.id,
        operation="restore",
        summary=f"Restore to v{version}",
    )

    # 如果被回滚的版本有 LLD Cell 明细，则将其复制到新版本
    old_cells = (
        db.query(SitePlanningCell)
        .filter(SitePlanningCell.planning_id == planning.id)
        .all()
    )
    if old_cells:
        for oc in old_cells:
            db.add(SitePlanningCell(planning_id=new_planning.id, **_cell_kwargs_from_model(oc)))
        db.commit()

    return await get_current_planning(site_id, db, current_user)


@router.post("/{site_id}/planning/upload", response_model=PlanningImportReport)
async def upload_planning(
    site_id: int,
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    # 生命周期门禁
    _ensure_site_plannable(site)

    content = await file.read()
    errors: List[str] = []
    parsed: Optional[SitePlanningBase] = None

    try:
        if file.filename.lower().endswith((".xlsx", ".xls")):
            excel = pd.ExcelFile(io.BytesIO(content))
            # expected sheets
            sheets = excel.sheet_names
            if "Summary" not in sheets:
                errors.append("缺少工作表: Summary")
            summary_df = excel.parse("Summary") if "Summary" in sheets else pd.DataFrame()
            sectors_df = excel.parse("Sectors") if "Sectors" in sheets else pd.DataFrame()
            ant_df = excel.parse("AntennaPorts") if "AntennaPorts" in sheets else pd.DataFrame()
            sw_df = excel.parse("SwitchPorts") if "SwitchPorts" in sheets else pd.DataFrame()

            # build planning
            bands = []
            sector_count = 0
            notes = None
            if not summary_df.empty:
                # assume first row
                row = summary_df.iloc[0].to_dict()
                bands_val = row.get("Bands") or ""
                bands = [b.strip() for b in str(bands_val).split(",") if str(b).strip()]
                sector_count = int(row.get("SectorCount") or 0)
                notes = row.get("Notes")

            sectors = []
            if not sectors_df.empty:
                for _, r in sectors_df.iterrows():
                    sectors.append({
                        "sector_index": int(r.get("SectorIndex") or 0),
                        "azimuth_deg": float(r.get("AzimuthDeg") or 0),
                        "downtilt_deg": float(r.get("DowntiltDeg") or 0),
                        "bands": [b.strip() for b in str(r.get("Bands") or "").split(",") if str(b).strip()],
                    })

            antenna_ports = []
            if not ant_df.empty:
                for _, r in ant_df.iterrows():
                    antenna_ports.append({
                        "port_label": str(r.get("PortLabel") or ""),
                        "sector_index": int(r.get("SectorIndex") or 0),
                        "band": (str(r.get("Band")) if not pd.isna(r.get("Band")) else None),
                        "mimo_chain": (str(r.get("MIMOChain")) if not pd.isna(r.get("MIMOChain")) else None),
                        "remarks": (str(r.get("Remarks")) if not pd.isna(r.get("Remarks")) else None),
                    })

            switch_ports = []
            if not sw_df.empty:
                for _, r in sw_df.iterrows():
                    vlans = str(r.get("VLANs") or "")
                    vlan_ids = [int(x) for x in vlans.split(",") if str(x).strip().isdigit()]
                    switch_ports.append({
                        "port_no": str(r.get("PortNo") or ""),
                        "vlan_ids": vlan_ids,
                        "is_uplink": bool(r.get("IsUplink") == True or str(r.get("IsUplink")).lower() in ("true", "1", "yes")),
                        "poe": bool(r.get("POE") == True or str(r.get("POE")).lower() in ("true", "1", "yes")),
                        "description": (str(r.get("Desc")) if not pd.isna(r.get("Desc")) else None),
                    })

            parsed = SitePlanningBase(
                bands=bands,
                sector_count=sector_count,
                notes=notes,
                sectors=sectors,
                antenna_ports=antenna_ports,
                switch_ports=switch_ports,
            )

        elif file.filename.lower().endswith(".json"):
            import json
            obj = json.loads(content.decode("utf-8"))
            parsed = SitePlanningBase(**obj)
        else:
            errors.append("仅支持 Excel(.xlsx/.xls) 或 JSON 文件")

    except Exception as e:
        errors.append(f"解析失败: {str(e)}")

    if errors:
        return PlanningImportReport(dry_run=dry_run, success=False, errors=errors, warnings=[], parsed=None)

    if dry_run:
        return PlanningImportReport(dry_run=True, success=True, errors=[], warnings=[], parsed=parsed)

    # persist as new version
    planning = _create_new_version(db, site_id, parsed, current_user.id, operation="import", summary=f"Import from {file.filename}")
    return PlanningImportReport(dry_run=False, success=True, errors=[], warnings=[], parsed=parsed)


@router.get("/planning/import-template")
async def download_import_template():
    # produce Excel with 4 sheets
    from fastapi.responses import StreamingResponse

    summary_df = pd.DataFrame([
        {"SiteCode": "SITE001", "Bands": "n41,n78", "SectorCount": 3, "Notes": "example"}
    ])
    sectors_df = pd.DataFrame([
        {"SectorIndex": 1, "AzimuthDeg": 0, "DowntiltDeg": 5, "Bands": "n41"},
        {"SectorIndex": 2, "AzimuthDeg": 120, "DowntiltDeg": 5, "Bands": "n41"},
        {"SectorIndex": 3, "AzimuthDeg": 240, "DowntiltDeg": 5, "Bands": "n41"},
    ])
    ant_df = pd.DataFrame([
        {"PortLabel": "ANT1", "SectorIndex": 1, "Band": "n41", "MIMOChain": "4x4", "Remarks": ""}
    ])
    sw_df = pd.DataFrame([
        {"PortNo": "GE1", "VLANs": "201,202", "IsUplink": True, "POE": False, "Desc": "Uplink"}
    ])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        sectors_df.to_excel(writer, sheet_name="Sectors", index=False)
        ant_df.to_excel(writer, sheet_name="AntennaPorts", index=False)
        sw_df.to_excel(writer, sheet_name="SwitchPorts", index=False)
    output.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=site_planning_template.xlsx"
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/{site_id}/planning/logs", response_model=List[PlanningChangeLogItem])
async def get_change_logs(
    site_id: int,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    logs = (
        db.query(PlanningChangeLog)
        .filter(PlanningChangeLog.site_id == site_id)
        .order_by(PlanningChangeLog.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    items: List[PlanningChangeLogItem] = []

    for l in logs:
        diff: Dict[str, Any] = l.diff or {}

        # 针对 LLD 相关操作，如果当初未写入 cell_changes，则在查询时按需补齐一次，
        # 以便前端“查看详情”能看到字段级别的 Cell 差异。
        try:
            if (l.operation or "").startswith("lld") and not diff.get("cell_changes"):
                planning = l.planning
                if planning:
                    current_plan = planning
                    prev_plan = (
                        db.query(SitePlanning)
                        .filter(
                            SitePlanning.site_id == l.site_id,
                            SitePlanning.version < current_plan.version,
                        )
                        .order_by(SitePlanning.version.desc())
                        .first()
                    )

                    old_cells: List[SitePlanningCell] = []
                    if prev_plan:
                        old_cells = (
                            db.query(SitePlanningCell)
                            .filter(SitePlanningCell.planning_id == prev_plan.id)
                            .all()
                        )

                    new_cells_db = (
                        db.query(SitePlanningCell)
                        .filter(SitePlanningCell.planning_id == current_plan.id)
                        .all()
                    )
                    new_cells: List[dict] = []
                    for c in new_cells_db:
                        d = {col.name: getattr(c, col.name) for col in SitePlanningCell.__table__.columns}
                        new_cells.append(d)

                    cell_diff = _compute_lld_cells_diff(old_cells, new_cells)
                    if cell_diff:
                        existing_cf = set(diff.get("changed_fields") or [])
                        for f in cell_diff.get("changed_fields") or []:
                            existing_cf.add(f)
                        if existing_cf:
                            diff["changed_fields"] = sorted(existing_cf)

                        existing_cc = diff.get("cell_changes") or []
                        existing_cc.extend(cell_diff.get("cell_changes") or [])
                        if existing_cc:
                            diff["cell_changes"] = existing_cc
        except Exception:
            # 补充 diff 失败不应影响主流程
            pass

        items.append(
            PlanningChangeLogItem(
                id=l.id,
                operation=l.operation,
                actor_id=l.actor_id,
                actor_name=getattr(l.actor, "username", None),
                summary=l.summary,
                created_at=l.created_at.isoformat() if l.created_at else "",
                diff=diff,
            )
        )

    return items


@router.get("/planning/batch-template")
async def download_batch_template():
    from fastapi.responses import StreamingResponse
    # Multi-site sheets include SiteCode column in all detail sheets
    sites_df = pd.DataFrame([
        {
            "SiteCode": "SITE001",
            "SiteName": "样例站点A",
            "SiteType": "macro",
            "Province": "北京",
            "City": "北京",
            "District": "朝阳区",
            "Address": "某路1号",
            "Priority": "normal",
            "ContactPerson": "张三",
            "ContactPhone": "13800000000",
            "Bands": "n41,n78",
            "SectorCount": 3,
            "Notes": "示例备注A"
        },
        {
            "SiteCode": "SITE002",
            "SiteName": "样例站点B",
            "SiteType": "macro",
            "Province": "上海",
            "City": "上海",
            "District": "浦东新区",
            "Address": "某路2号",
            "Priority": "high",
            "ContactPerson": "李四",
            "ContactPhone": "13900000000",
            "Bands": "n41",
            "SectorCount": 3,
            "Notes": "示例备注B"
        },
    ])
    sectors_df = pd.DataFrame([
        {"SiteCode": "SITE001", "SectorIndex": 1, "AzimuthDeg": 0, "DowntiltDeg": 5, "Bands": "n41"},
        {"SiteCode": "SITE001", "SectorIndex": 2, "AzimuthDeg": 120, "DowntiltDeg": 5, "Bands": "n41"},
        {"SiteCode": "SITE001", "SectorIndex": 3, "AzimuthDeg": 240, "DowntiltDeg": 5, "Bands": "n41"},
        {"SiteCode": "SITE002", "SectorIndex": 1, "AzimuthDeg": 0, "DowntiltDeg": 6, "Bands": "n41"},
        {"SiteCode": "SITE002", "SectorIndex": 2, "AzimuthDeg": 120, "DowntiltDeg": 6, "Bands": "n41"},
        {"SiteCode": "SITE002", "SectorIndex": 3, "AzimuthDeg": 240, "DowntiltDeg": 6, "Bands": "n41"},
    ])
    ant_df = pd.DataFrame([
        {"SiteCode": "SITE001", "PortLabel": "ANT1", "SectorIndex": 1, "Band": "n41", "MIMOChain": "4x4", "Remarks": ""},
        {"SiteCode": "SITE002", "PortLabel": "ANT1", "SectorIndex": 1, "Band": "n41", "MIMOChain": "4x4", "Remarks": ""},
    ])
    sw_df = pd.DataFrame([
        {"SiteCode": "SITE001", "PortNo": "GE1", "VLANs": "201,202", "IsUplink": True, "POE": False, "Desc": "Uplink"},
        {"SiteCode": "SITE002", "PortNo": "GE1", "VLANs": "301,302", "IsUplink": True, "POE": False, "Desc": "Uplink"},
    ])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sites_df.to_excel(writer, sheet_name="Sites", index=False)
        sectors_df.to_excel(writer, sheet_name="Sectors", index=False)
        ant_df.to_excel(writer, sheet_name="AntennaPorts", index=False)
        sw_df.to_excel(writer, sheet_name="SwitchPorts", index=False)
    output.seek(0)
    headers = {
        "Content-Disposition": "attachment; filename=site_planning_batch_template.xlsx"
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.post("/planning/batch-upload", response_model=BatchImportReport)
async def batch_upload_planning(
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    content = await file.read()
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 Excel(.xlsx/.xls)")

    excel = pd.ExcelFile(io.BytesIO(content))
    required = ["Sites", "Sectors", "AntennaPorts", "SwitchPorts"]
    for name in required:
        if name not in excel.sheet_names:
            raise HTTPException(status_code=400, detail=f"缺少工作表: {name}")

    sites_df = excel.parse("Sites")
    sectors_df = excel.parse("Sectors")
    ant_df = excel.parse("AntennaPorts")
    sw_df = excel.parse("SwitchPorts")

    results: List[BatchPlanningResult] = []
    success_count = 0
    failed_count = 0

    site_codes = [str(x).strip() for x in sites_df.get("SiteCode", []) if str(x).strip()]
    unique_codes = list(dict.fromkeys(site_codes))

    for code in unique_codes:
        errors: List[str] = []
        warnings: List[str] = []
        try:
            site = db.query(Site).filter(Site.site_code == code).first()
            if not site:
                errors.append("站点不存在：请先在‘站点信息导入’创建站点并完成勘察")
                results.append(BatchPlanningResult(site_code=code, success=False, errors=errors))
                failed_count += 1
                continue
            # 生命周期门禁
            try:
                _ensure_site_plannable(site)
            except HTTPException as he:
                errors.append(str(he.detail))
                results.append(BatchPlanningResult(site_code=code, site_id=site.id, success=False, errors=errors))
                failed_count += 1
                continue

            srow = sites_df[sites_df["SiteCode"].astype(str) == code].iloc[0].to_dict()
            bands_val = srow.get("Bands") or ""
            bands = [b.strip() for b in str(bands_val).split(",") if str(b).strip()]
            sector_count = int(srow.get("SectorCount") or 0)
            notes = srow.get("Notes")

            sec_rows = sectors_df[sectors_df["SiteCode"].astype(str) == code]
            sectors = []
            for _, r in sec_rows.iterrows():
                sectors.append({
                    "sector_index": int(r.get("SectorIndex") or 0),
                    "azimuth_deg": float(r.get("AzimuthDeg") or 0),
                    "downtilt_deg": float(r.get("DowntiltDeg") or 0),
                    "bands": [b.strip() for b in str(r.get("Bands") or "").split(",") if str(b).strip()],
                })

            ant_rows = ant_df[ant_df["SiteCode"].astype(str) == code]
            ants = []
            for _, r in ant_rows.iterrows():
                ants.append({
                    "port_label": str(r.get("PortLabel") or ""),
                    "sector_index": int(r.get("SectorIndex") or 0),
                    "band": (str(r.get("Band")) if pd.notna(r.get("Band")) else None),
                    "mimo_chain": (str(r.get("MIMOChain")) if pd.notna(r.get("MIMOChain")) else None),
                    "remarks": (str(r.get("Remarks")) if pd.notna(r.get("Remarks")) else None),
                })

            sw_rows = sw_df[sw_df["SiteCode"].astype(str) == code]
            sps = []
            for _, r in sw_rows.iterrows():
                vlans = str(r.get("VLANs") or "")
                vlan_ids = [int(x) for x in vlans.split(",") if str(x).strip().isdigit()]
                sps.append({
                    "port_no": str(r.get("PortNo") or ""),
                    "vlan_ids": vlan_ids,
                    "is_uplink": bool(r.get("IsUplink") == True or str(r.get("IsUplink")).lower() in ("true", "1", "yes")),
                    "poe": bool(r.get("POE") == True or str(r.get("POE")).lower() in ("true", "1", "yes")),
                    "description": (str(r.get("Desc")) if pd.notna(r.get("Desc")) else None),
                })

            parsed = SitePlanningBase(
                bands=bands,
                sector_count=sector_count,
                notes=notes,
                sectors=sectors,
                antenna_ports=ants,
                switch_ports=sps,
            )

            if not dry_run:
                newp = _create_new_version(db, site.id, parsed, current_user.id, operation="import", summary=f"Batch import: {file.filename}")
                results.append(BatchPlanningResult(site_code=code, site_id=site.id, success=True, version_created=newp.version, warnings=warnings))
            else:
                results.append(BatchPlanningResult(site_code=code, site_id=(site.id if site else None), success=True, warnings=warnings))
            success_count += 1

        except Exception as e:
            errors.append(str(e))
            results.append(BatchPlanningResult(site_code=code, success=False, errors=errors))
            failed_count += 1

    return BatchImportReport(
        dry_run=dry_run,
        total_sites=len(unique_codes),
        success_count=success_count,
        failed_count=failed_count,
        results=results,
    )


@router.get("/planning/lld-batch-template")
async def download_lld_batch_template():
    """
    下载 LLD 规划导入模板。

    返回仓库根目录下的模板 `LLD templateV1.0.xlsx`（Sheet: 4G/5G）。
    """
    from fastapi.responses import StreamingResponse
    from pathlib import Path

    base_dir = Path(__file__).resolve().parents[3]
    xlsx_path = base_dir / "LLD templateV1.0.xlsx"
    if not xlsx_path.exists():
        raise HTTPException(
            status_code=500,
            detail="LLD 模板文件缺失：请在系统部署目录中放置 LLD templateV1.0.xlsx",
        )

    data = xlsx_path.read_bytes()
    output = io.BytesIO(data)
    output.seek(0)
    headers = {
        "Content-Disposition": "attachment; filename=site_planning_lld_template.xlsx"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/planning/lld-batch-upload", response_model=BatchImportReport)
async def lld_batch_upload_planning(
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    基于 LLD Excel 批量导入站点规划。

    - 仅支持新模板：Sheet 为 4G/5G（频段从行内 BAND/Band 列读取）

    使用 SiteCode 作为 SITE ID，匹配 sites.site_code。
    """
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    content = await file.read()
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 Excel(.xlsx/.xls)")

    try:
        excel = pd.ExcelFile(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {e}")

    rows_by_tower = _collect_lld_rows_by_tower(excel)
    if not rows_by_tower:
        raise HTTPException(
            status_code=400,
            detail="未在 LLD 文件中找到有效的规划数据（请检查是否包含 4G/5G Sheet，且 SiteCode 与 4G: Sector ID Local ID / 5G: Sector ID 列存在且非空）",
        )

    results: List[BatchPlanningResult] = []
    success_count = 0
    failed_count = 0

    for site_code, metas in rows_by_tower.items():
        errors: List[str] = []
        warnings: List[str] = []
        try:
            site = db.query(Site).filter(Site.site_code == site_code).first()
            if not site:
                errors.append("站点不存在：请先在‘站点信息导入’创建站点并完成勘察")
                results.append(BatchPlanningResult(site_code=site_code, success=False, errors=errors))
                failed_count += 1
                continue

            # 生命周期门禁
            try:
                _ensure_site_plannable(site)
            except HTTPException as he:
                errors.append(str(he.detail))
                results.append(BatchPlanningResult(site_code=site_code, site_id=site.id, success=False, errors=errors))
                failed_count += 1
                continue

            # 构造 Cell 明细
            cell_dicts: List[dict] = []
            for meta in metas:
                cell = _build_cell_dict_from_row(site.id, meta.get("tower_id"), meta)
                if cell.get("local_cell_id") is None:
                    # 理论上上游已经过滤过，这里再次防御
                    continue
                cell_dicts.append(cell)

            if not cell_dicts:
                errors.append("该站点在 LLD 中没有有效的小区行（4G: Sector ID Local ID / 5G: Sector ID 为空）")
                results.append(BatchPlanningResult(site_code=site_code, site_id=site.id, success=False, errors=errors))
                failed_count += 1
                continue

            lte_cell_count = sum(1 for c in cell_dicts if c.get("rat") == "LTE")
            nr_cell_count = sum(1 for c in cell_dicts if c.get("rat") == "NR")
            bands = sorted({c.get("band_code") for c in cell_dicts if c.get("band_code")})

            has_active_opening_wo = _has_active_opening_work_order(db, site.id)
            edit_policy = _compute_lld_edit_policy(site, has_active_opening_wo, current_user)

            # 用于受限策略校验/预览的 diff
            current_planning = _get_current_planning(db, site.id)
            old_cells: List[SitePlanningCell] = []
            if current_planning:
                old_cells = (
                    db.query(SitePlanningCell)
                    .filter(SitePlanningCell.planning_id == current_planning.id)
                    .all()
                )
            preview_diff = _compute_lld_cells_diff(old_cells, cell_dicts)

            conflicts = _find_lld_policy_conflicts_for_import(edit_policy, preview_diff)
            if conflicts:
                errors.extend(_build_lld_policy_conflict_errors(edit_policy, conflicts))
                results.append(
                    BatchPlanningResult(
                        site_code=site_code,
                        site_id=site.id,
                        success=False,
                        errors=errors,
                        warnings=warnings,
                        lte_cell_count=lte_cell_count,
                        nr_cell_count=nr_cell_count,
                        bands=bands,
                        preview_diff=preview_diff or None,
                    )
                )
                failed_count += 1
                continue

            if dry_run:
                # 试运行：不落库，但返回更详细的 Cell 变更预览，方便前端展示
                results.append(
                    BatchPlanningResult(
                        site_code=site_code,
                        site_id=site.id,
                        success=True,
                        warnings=warnings,
                        lte_cell_count=lte_cell_count,
                        nr_cell_count=nr_cell_count,
                        bands=bands,
                        preview_diff=preview_diff or None,
                    )
                )
                success_count += 1
                continue

            # 构造并持久化新的规划版本
            base = _build_planning_from_cells(site.id, cell_dicts)
            planning = _create_new_version(
                db,
                site.id,
                base,
                current_user.id,
                operation="lld_import",
                summary=f"LLD batch import: {file.filename}",
            )

            # 将 cells 与规划版本关联后落库
            for c in cell_dicts:
                db.add(SitePlanningCell(planning_id=planning.id, **c))
            db.commit()

            results.append(
                BatchPlanningResult(
                    site_code=site_code,
                    site_id=site.id,
                    success=True,
                    version_created=planning.version,
                    warnings=warnings,
                    lte_cell_count=lte_cell_count,
                    nr_cell_count=nr_cell_count,
                    bands=bands,
                )
            )
            success_count += 1
        except Exception as e:
            errors.append(str(e))
            results.append(BatchPlanningResult(site_code=site_code, success=False, errors=errors))
            failed_count += 1

    return BatchImportReport(
        dry_run=dry_run,
        total_sites=len(rows_by_tower),
        success_count=success_count,
        failed_count=failed_count,
        results=results,
    )


@router.post("/{site_id}/planning/lld-upload", response_model=BatchPlanningResult)
async def lld_upload_planning_for_site(
    site_id: int,
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    单站 LLD 导入：只处理 LLD 文件中 SiteCode 与该站点 site_code 匹配的行。
    返回 BatchPlanningResult 以复用前端展示逻辑。
    """
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    _ensure_site_plannable(site)

    content = await file.read()
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 Excel(.xlsx/.xls)")

    try:
        excel = pd.ExcelFile(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {e}")

    rows_by_tower = _collect_lld_rows_by_tower(excel)
    metas = rows_by_tower.get(site.site_code) or []
    if not metas:
        raise HTTPException(status_code=400, detail=f"LLD 文件中未找到与站点编码 {site.site_code} 匹配的 SiteCode 行")

    errors: List[str] = []
    warnings: List[str] = []

    # 记录导入前的 Cell 列表，用于后续生成 diff
    current_planning = _get_current_planning(db, site.id)
    old_cells: List[SitePlanningCell] = []
    if current_planning:
        old_cells = (
            db.query(SitePlanningCell)
            .filter(
                SitePlanningCell.site_id == site.id,
                SitePlanningCell.planning_id == current_planning.id,
            )
            .all()
        )

    cell_dicts: List[dict] = []
    for meta in metas:
        cell = _build_cell_dict_from_row(site.id, meta.get("tower_id"), meta)
        if cell.get("local_cell_id") is None:
            continue
        cell_dicts.append(cell)

    if not cell_dicts:
        errors.append("该站点在 LLD 中没有有效的小区行（4G: Sector ID Local ID / 5G: Sector ID 为空）")
        return BatchPlanningResult(site_code=site.site_code, site_id=site.id, success=False, errors=errors)

    lte_cell_count = sum(1 for c in cell_dicts if c.get("rat") == "LTE")
    nr_cell_count = sum(1 for c in cell_dicts if c.get("rat") == "NR")
    bands = sorted({c.get("band_code") for c in cell_dicts if c.get("band_code")})

    has_active_opening_wo = _has_active_opening_work_order(db, site.id)
    edit_policy = _compute_lld_edit_policy(site, has_active_opening_wo, current_user)

    # 预先计算 diff（用于受限策略校验与 dry_run 预览）
    cell_diff = _compute_lld_cells_diff(old_cells, cell_dicts)

    conflicts = _find_lld_policy_conflicts_for_import(edit_policy, cell_diff)
    if conflicts:
        errors.extend(_build_lld_policy_conflict_errors(edit_policy, conflicts))
        return BatchPlanningResult(
            site_code=site.site_code,
            site_id=site.id,
            success=False,
            errors=errors,
            warnings=warnings,
            lte_cell_count=lte_cell_count,
            nr_cell_count=nr_cell_count,
            bands=bands,
            preview_diff=cell_diff or None,
        )

    if dry_run:
        return BatchPlanningResult(
            site_code=site.site_code,
            site_id=site.id,
            success=True,
            warnings=warnings,
            lte_cell_count=lte_cell_count,
            nr_cell_count=nr_cell_count,
            bands=bands,
            preview_diff=cell_diff or None,
        )

    base = _build_planning_from_cells(site.id, cell_dicts)
    planning = _create_new_version(
        db,
        site.id,
        base,
        current_user.id,
        operation="lld_import",
        summary=f"LLD single-site import: {file.filename}",
    )
    for c in cell_dicts:
        db.add(SitePlanningCell(planning_id=planning.id, **c))
    db.commit()

    # 基于导入前后的 Cell 列表，将更细粒度的 diff（字段级别变更）合并写入日志
    try:
        _merge_lld_cell_diff_into_planning_log(db, planning.id, cell_diff)
    except Exception:
        # diff 生成失败不影响主流程
        pass

    return BatchPlanningResult(
        site_code=site.site_code,
        site_id=site.id,
        success=True,
        version_created=planning.version,
        warnings=warnings,
        lte_cell_count=lte_cell_count,
        nr_cell_count=nr_cell_count,
        bands=bands,
    )


@router.get("/{site_id}/planning/lld", response_model=SitePlanningLldResponse)
async def get_lld_planning(
    site_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    获取当前版本的 LLD 规划数据：包括 SitePlanning 概要 + Cell 明细 + 汇总信息。
    """
    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")

    has_active_opening_wo = _has_active_opening_work_order(db, site_id)
    edit_policy = _compute_lld_edit_policy(site, has_active_opening_wo, current_user)

    planning = _get_current_planning(db, site_id)
    if not planning:
        raise HTTPException(status_code=404, detail="Planning not found for this site")

    cells = (
        db.query(SitePlanningCell)
        .filter(
            SitePlanningCell.site_id == site_id,
            SitePlanningCell.planning_id == planning.id,
        )
        .order_by(SitePlanningCell.rat.asc(), SitePlanningCell.band_code.asc(), SitePlanningCell.local_cell_id.asc())
        .all()
    )

    summary = _build_lld_summary_from_cells(planning, cells)
    # 直接复用已有的 get_current_planning 构建概要
    planning_resp = await get_current_planning(site_id, db, current_user)

    cell_models = [PlanningCell.model_validate(c, from_attributes=True) for c in cells]

    return SitePlanningLldResponse(
        planning=planning_resp,
        cells=cell_models,
        summary=summary,
        edit_policy=edit_policy,
    )


@router.put("/{site_id}/planning/lld", response_model=SitePlanningLldResponse)
async def update_lld_planning(
    site_id: int,
    base_version: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    批量更新 LLD 规划（基于现有规划创建新版本）。
    主要用于前端编辑后保存整个规划状态，保持版本管理。
    """
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    _ensure_site_plannable(site)

    current = _get_current_planning(db, site_id)
    if not current:
        raise HTTPException(status_code=404, detail="No existing planning found for this site")

    # 乐观锁检查
    if current.version != base_version:
        raise HTTPException(
            status_code=409,
            detail=f"Version conflict: current={current.version}, base={base_version}"
        )

    # 创建新版本（当前无变更，只是版本递增）
    old_snapshot = _snapshot(current)
    data = SitePlanningBase(
        bands=current.bands or [],
        sector_count=current.sector_count or 0,
        notes=current.notes,
        sectors=[
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in (current.sectors or [])
        ],
        antenna_ports=[
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in (current.antenna_ports or [])
        ],
        switch_ports=[
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in (current.switch_ports or [])
        ],
    )

    new_planning = _create_new_version(
        db,
        site_id,
        data,
        current_user.id,
        operation="lld_update",
        summary="Manual LLD planning update",
    )

    # 复制当前版本的 Cell 明细到新版本，避免出现“新版本为空”的风险
    old_cells = (
        db.query(SitePlanningCell)
        .filter(SitePlanningCell.planning_id == current.id)
        .all()
    )
    for oc in old_cells:
        db.add(SitePlanningCell(planning_id=new_planning.id, **_cell_kwargs_from_model(oc)))
    db.flush()

    all_cells = db.query(SitePlanningCell).filter(SitePlanningCell.planning_id == new_planning.id).all()
    _sync_planning_from_cells(db, new_planning, all_cells)
    db.commit()

    return await get_lld_planning(site_id, db, current_user)


@router.post("/{site_id}/planning/lld/cells", response_model=PlanningCell)
async def create_lld_cell(
    site_id: int,
    cell_data: PlanningCellCreate,
    base_version: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    添加新的 LLD Cell。
    会自动创建新规划版本并更新基础规划数据。
    """
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    _ensure_site_plannable(site)

    has_active_opening_wo = _has_active_opening_work_order(db, site_id)
    edit_policy = _compute_lld_edit_policy(site, has_active_opening_wo, current_user)
    if edit_policy.mode != "full":
        raise HTTPException(
            status_code=409,
            detail={
                "code": "LLD_EDIT_MODE_RESTRICTED",
                "message": "当前站点处于受限编辑状态，禁止新增 Cell",
                "reason": edit_policy.reason,
                "mode": edit_policy.mode,
                "locked_fields": edit_policy.locked_fields,
            },
        )

    current = _get_current_planning(db, site_id)
    if not current:
        raise HTTPException(status_code=404, detail="No existing planning found for this site")

    # 乐观锁检查
    if current.version != base_version:
        raise HTTPException(
            status_code=409,
            detail=f"Version conflict: current={current.version}, base={base_version}"
        )

    # 验证Cell数据
    cell_dict = cell_data.model_dump()
    validation_errors = _validate_cell_data(cell_dict, "create")
    if validation_errors:
        raise HTTPException(status_code=400, detail={"errors": validation_errors})

    # 检查Cell冲突
    conflicts = _check_cell_conflicts(db, site_id, cell_dict)
    if conflicts:
        raise HTTPException(status_code=409, detail={"conflicts": conflicts})

    old_cells = (
        db.query(SitePlanningCell)
        .filter(SitePlanningCell.planning_id == current.id)
        .all()
    )

    # 创建新规划版本
    data = SitePlanningBase(
        bands=current.bands or [],
        sector_count=current.sector_count or 0,
        notes=current.notes,
        sectors=[
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in (current.sectors or [])
        ],
        antenna_ports=[
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in (current.antenna_ports or [])
        ],
        switch_ports=[
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in (current.switch_ports or [])
        ],
    )

    new_planning = _create_new_version(
        db,
        site_id,
        data,
        current_user.id,
        operation="lld_add_cell",
        summary=f"Add new {cell_data.rat} cell: {cell_data.band_code}",
    )

    # 复制旧版本的所有 Cell 到新版本
    for oc in old_cells:
        db.add(SitePlanningCell(planning_id=new_planning.id, **_cell_kwargs_from_model(oc)))

    # 创建新的 Cell
    cell_payload = cell_dict.copy()
    if 'sheet_name' not in cell_payload or not cell_payload['sheet_name']:
        cell_payload['sheet_name'] = f"{cell_payload.get('rat', 'UNKNOWN')}-{cell_payload.get('band_code', 'UNKNOWN')}"
    # TOWER ID 允许手动填写；未填写/空白则默认使用站点编码
    tower_id = cell_payload.get("tower_id")
    if tower_id is None or (isinstance(tower_id, str) and not tower_id.strip()):
        cell_payload["tower_id"] = site.site_code
    else:
        cell_payload["tower_id"] = str(tower_id).strip()
    # 补齐站点信息字段（便于导出/展示）
    if not cell_payload.get("site_information"):
        cell_payload["site_information"] = site.site_code
    if not cell_payload.get("site_name"):
        cell_payload["site_name"] = site.site_name

    cell = SitePlanningCell(
        planning_id=new_planning.id,
        site_id=site_id,
        **cell_payload
    )
    db.add(cell)
    db.flush()

    # 同步基础规划数据
    all_cells = db.query(SitePlanningCell).filter(SitePlanningCell.planning_id == new_planning.id).all()
    _sync_planning_from_cells(db, new_planning, all_cells)
    db.commit()
    db.refresh(cell)

    # 合并 Cell 级 diff 到日志，便于前端展示
    try:
        new_cells_dicts = [_cell_kwargs_from_model(oc) for oc in old_cells] + [_cell_kwargs_from_model(cell)]
        cell_diff = _compute_lld_cells_diff(old_cells, new_cells_dicts)
        _merge_lld_cell_diff_into_planning_log(db, new_planning.id, cell_diff)
    except Exception:
        pass

    return PlanningCell.model_validate(cell, from_attributes=True)


@router.put("/{site_id}/planning/lld/cells/{cell_id}", response_model=PlanningCell)
async def update_lld_cell(
    site_id: int,
    cell_id: int,
    cell_data: PlanningCellUpdate,
    base_version: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    更新指定的 LLD Cell。
    会自动创建新规划版本并复制其他 Cell 到新版本。
    """
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    _ensure_site_plannable(site)

    has_active_opening_wo = _has_active_opening_work_order(db, site_id)
    edit_policy = _compute_lld_edit_policy(site, has_active_opening_wo, current_user)

    current = _get_current_planning(db, site_id)
    if not current:
        raise HTTPException(status_code=404, detail="No existing planning found for this site")

    # 乐观锁检查
    if current.version != base_version:
        raise HTTPException(
            status_code=409,
            detail=f"Version conflict: current={current.version}, base={base_version}"
        )

    # 查找要更新的 Cell
    existing_cell = (
        db.query(SitePlanningCell)
        .filter(
            SitePlanningCell.id == cell_id,
            SitePlanningCell.site_id == site_id,
            SitePlanningCell.planning_id == current.id,
        )
        .first()
    )
    if not existing_cell:
        raise HTTPException(status_code=404, detail="Cell not found")

    # 验证更新数据，并预先构建 Cell 级字段差异，便于写入变更日志
    update_dict = cell_data.model_dump(exclude_unset=True)
    field_changes: List[Dict[str, Any]] = []
    if update_dict:
        if edit_policy.mode == "limited":
            attempted_locked = sorted({k for k in update_dict.keys() if k in LLD_LOCKED_FIELDS})
            if attempted_locked:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "LLD_EDIT_LOCKED_FIELDS",
                        "message": "当前站点处于受限编辑状态，禁止修改关键字段",
                        "reason": edit_policy.reason,
                        "mode": edit_policy.mode,
                        "locked_fields": edit_policy.locked_fields,
                        "attempted_fields": attempted_locked,
                    },
                )

        # 关键字段不允许被清空（允许在 full 模式下修改为新的非空值）
        required_errors: List[str] = []
        for f, msg in [
            ("rat", "RAT不能为空"),
            ("band_code", "频段代码不能为空"),
            ("local_cell_id", "Local Cell ID不能为空"),
        ]:
            if f in update_dict:
                v = update_dict.get(f)
                if v is None or (isinstance(v, str) and not str(v).strip()):
                    required_errors.append(msg)

        # tower_id 允许修改；若传空/空白则按“默认填充站点编码”处理
        if "tower_id" in update_dict:
            v = update_dict.get("tower_id")
            if v is None or (isinstance(v, str) and not str(v).strip()):
                update_dict["tower_id"] = site.site_code
            else:
                update_dict["tower_id"] = str(v).strip()

        validation_errors = _validate_cell_data(update_dict, "update") + required_errors
        if validation_errors:
            raise HTTPException(status_code=400, detail={"errors": validation_errors})

        # 检查更新后的数据是否会造成冲突
        merged_data = existing_cell.__dict__.copy()
        merged_data.update(update_dict)
        conflicts = _check_cell_conflicts(db, site_id, merged_data, exclude_cell_id=cell_id)
        if conflicts:
            raise HTTPException(status_code=409, detail={"conflicts": conflicts})

        # 记录字段级别变更（例如 TAC 从 A 改成 B），用于后续写入 PlanningChangeLog.diff
        for field, new_value in update_dict.items():
            old_value = getattr(existing_cell, field, None)
            if old_value != new_value:
                field_changes.append(
                    {
                        "field": field,
                        "old": old_value,
                        "new": new_value,
                    }
                )

    # 创建新规划版本
    old_snapshot = _snapshot(current)
    data = SitePlanningBase(
        bands=current.bands or [],
        sector_count=current.sector_count or 0,
        notes=current.notes,
        sectors=[
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in (current.sectors or [])
        ],
        antenna_ports=[
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in (current.antenna_ports or [])
        ],
        switch_ports=[
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in (current.switch_ports or [])
        ],
    )

    new_planning = _create_new_version(
        db,
        site_id,
        data,
        current_user.id,
        operation="lld_update_cell",
        summary=f"Update {existing_cell.rat} cell: {existing_cell.band_code}",
    )

    # 复制旧版本的所有 Cell 到新版本，并在目标 Cell 上应用更新（不破坏旧版本历史数据）
    old_cells = (
        db.query(SitePlanningCell)
        .filter(SitePlanningCell.planning_id == current.id)
        .all()
    )

    new_cells_dicts: List[dict] = []
    updated_cell_obj: Optional[SitePlanningCell] = None

    for oc in old_cells:
        kwargs = _cell_kwargs_from_model(oc)
        if oc.id == cell_id:
            kwargs.update(update_dict)
        new_cells_dicts.append(kwargs)

        new_cell = SitePlanningCell(planning_id=new_planning.id, **kwargs)
        db.add(new_cell)
        if oc.id == cell_id:
            updated_cell_obj = new_cell

    db.flush()

    # 同步基础规划数据
    all_cells = db.query(SitePlanningCell).filter(SitePlanningCell.planning_id == new_planning.id).all()
    _sync_planning_from_cells(db, new_planning, all_cells)
    db.commit()

    # 合并 Cell 级 diff 到日志，便于前端展示（包含增删/关键字段变更等场景）
    try:
        cell_diff = _compute_lld_cells_diff(old_cells, new_cells_dicts)
        _merge_lld_cell_diff_into_planning_log(db, new_planning.id, cell_diff)
    except Exception:
        pass

    if not updated_cell_obj:
        raise HTTPException(status_code=500, detail="更新失败：未生成新的 Cell 记录")
    db.refresh(updated_cell_obj)
    return PlanningCell.model_validate(updated_cell_obj, from_attributes=True)


@router.delete("/{site_id}/planning/lld/cells/{cell_id}")
async def delete_lld_cell(
    site_id: int,
    cell_id: int,
    base_version: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    删除指定的 LLD Cell。
    会自动创建新规划版本并复制其他 Cell 到新版本。
    """
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    _ensure_site_plannable(site)

    has_active_opening_wo = _has_active_opening_work_order(db, site_id)
    edit_policy = _compute_lld_edit_policy(site, has_active_opening_wo, current_user)
    if edit_policy.mode != "full":
        raise HTTPException(
            status_code=409,
            detail={
                "code": "LLD_EDIT_MODE_RESTRICTED",
                "message": "当前站点处于受限编辑状态，禁止删除 Cell",
                "reason": edit_policy.reason,
                "mode": edit_policy.mode,
                "locked_fields": edit_policy.locked_fields,
            },
        )

    current = _get_current_planning(db, site_id)
    if not current:
        raise HTTPException(status_code=404, detail="No existing planning found for this site")

    # 乐观锁检查
    if current.version != base_version:
        raise HTTPException(
            status_code=409,
            detail=f"Version conflict: current={current.version}, base={base_version}"
        )

    # 查找要删除的 Cell
    existing_cell = (
        db.query(SitePlanningCell)
        .filter(
            SitePlanningCell.id == cell_id,
            SitePlanningCell.site_id == site_id,
            SitePlanningCell.planning_id == current.id,
        )
        .first()
    )
    if not existing_cell:
        raise HTTPException(status_code=404, detail="Cell not found")

    # 创建新规划版本
    old_snapshot = _snapshot(current)
    data = SitePlanningBase(
        bands=current.bands or [],
        sector_count=current.sector_count or 0,
        notes=current.notes,
        sectors=[
            {
                "sector_index": s.sector_index,
                "azimuth_deg": s.azimuth_deg,
                "downtilt_deg": s.downtilt_deg,
                "bands": s.bands or [],
            }
            for s in (current.sectors or [])
        ],
        antenna_ports=[
            {
                "port_label": p.port_label,
                "sector_index": p.sector_index,
                "band": p.band,
                "mimo_chain": p.mimo_chain,
                "remarks": p.remarks,
            }
            for p in (current.antenna_ports or [])
        ],
        switch_ports=[
            {
                "port_no": sp.port_no,
                "vlan_ids": sp.vlan_ids or [],
                "is_uplink": sp.is_uplink,
                "poe": sp.poe,
                "description": sp.description,
            }
            for sp in (current.switch_ports or [])
        ],
    )

    new_planning = _create_new_version(
        db,
        site_id,
        data,
        current_user.id,
        operation="lld_delete_cell",
        summary=f"Delete {existing_cell.rat} cell: {existing_cell.band_code}",
    )

    # 复制除要删除的 Cell 之外的所有 Cell
    old_cells = (
        db.query(SitePlanningCell)
        .filter(SitePlanningCell.planning_id == current.id)
        .all()
    )
    for old_cell in old_cells:
        if old_cell.id != cell_id:
            db.add(SitePlanningCell(planning_id=new_planning.id, **_cell_kwargs_from_model(old_cell)))

    db.commit()

    # 同步基础规划数据
    all_cells = db.query(SitePlanningCell).filter(SitePlanningCell.planning_id == new_planning.id).all()
    _sync_planning_from_cells(db, new_planning, all_cells)
    db.commit()

    return {"message": "Cell deleted successfully", "deleted_cell_id": cell_id}


def _build_lld_planning_query(
    db: Session,
    status: Optional[str],
    band_list: List[str],
    start_time: Optional[datetime],
    end_time: Optional[datetime],
):
    query = (
        db.query(SitePlanning, Site)
        .join(Site, SitePlanning.site_id == Site.id)
        .filter(SitePlanning.is_current == True)
    )

    if status:
        query = query.filter(Site.status == status)
    if start_time:
        query = query.filter(SitePlanning.updated_at >= start_time)
    if end_time:
        query = query.filter(SitePlanning.updated_at <= end_time)

    if band_list:
        band_subq = (
            db.query(SitePlanningCell.planning_id)
            .filter(SitePlanningCell.band_code.in_(band_list))
            .distinct()
            .subquery()
        )
        query = query.filter(SitePlanning.id.in_(band_subq))

    return query


def _collect_lld_planning_stats(
    db: Session,
    planning_ids: List[int],
) -> Tuple[Dict[int, Any], Dict[int, List[str]]]:
    stats_map: Dict[int, Any] = {}
    bands_map: Dict[int, List[str]] = {}

    if not planning_ids:
        return stats_map, bands_map

    stats_rows = (
        db.query(
            SitePlanningCell.planning_id.label("planning_id"),
            func.sum(case((SitePlanningCell.rat == "LTE", 1), else_=0)).label("lte_count"),
            func.sum(case((SitePlanningCell.rat == "NR", 1), else_=0)).label("nr_count"),
            func.min(SitePlanningCell.mechanical_downtilt_deg).label("mechanical_min"),
            func.max(SitePlanningCell.mechanical_downtilt_deg).label("mechanical_max"),
            func.min(SitePlanningCell.electrical_downtilt_deg).label("electrical_min"),
            func.max(SitePlanningCell.electrical_downtilt_deg).label("electrical_max"),
        )
        .filter(SitePlanningCell.planning_id.in_(planning_ids))
        .group_by(SitePlanningCell.planning_id)
        .all()
    )
    for row in stats_rows:
        stats_map[row.planning_id] = row

    band_rows = (
        db.query(SitePlanningCell.planning_id, SitePlanningCell.band_code)
        .filter(SitePlanningCell.planning_id.in_(planning_ids))
        .distinct()
        .all()
    )
    for pid, b in band_rows:
        if not b:
            continue
        bands_map.setdefault(pid, []).append(str(b))

    # 去重并排序
    for pid, items in bands_map.items():
        bands_map[pid] = sorted(list(dict.fromkeys(items)))

    return stats_map, bands_map


def _fetch_lld_planning_summary(
    db: Session,
    status: Optional[str],
    band: Optional[str],
    start_time: Optional[datetime],
    end_time: Optional[datetime],
    skip: int = 0,
    limit: Optional[int] = 50,
) -> Tuple[List[LldPlanningSummaryItem], int]:
    band_list = _parse_band_list(band)
    query = _build_lld_planning_query(db, status, band_list, start_time, end_time)
    total = query.count()

    if limit is not None:
        rows = (
            query.order_by(SitePlanning.updated_at.desc(), SitePlanning.id.desc())
            .offset(skip)
            .limit(limit)
            .all()
        )
    else:
        rows = query.order_by(SitePlanning.updated_at.desc(), SitePlanning.id.desc()).all()

    planning_ids = [planning.id for planning, _ in rows]
    stats_map, bands_map = _collect_lld_planning_stats(db, planning_ids)

    items: List[LldPlanningSummaryItem] = []
    for planning, site in rows:
        stats = stats_map.get(planning.id)
        bands = bands_map.get(planning.id) or planning.bands or []
        if isinstance(bands, str):
            bands = _parse_band_list(bands)

        items.append(
            LldPlanningSummaryItem(
                site_id=site.id,
                site_code=site.site_code,
                site_name=site.site_name,
                site_type=site.site_type,
                province=site.province,
                city=site.city,
                district=site.district,
                status=site.status,
                planning_id=planning.id,
                planning_version=planning.version,
                planning_created_at=planning.created_at,
                planning_updated_at=planning.updated_at,
                planning_notes=planning.notes,
                bands=bands,
                sector_count=planning.sector_count or 0,
                lte_cell_count=int(getattr(stats, "lte_count", 0) or 0),
                nr_cell_count=int(getattr(stats, "nr_count", 0) or 0),
                mechanical_downtilt_min=getattr(stats, "mechanical_min", None),
                mechanical_downtilt_max=getattr(stats, "mechanical_max", None),
                electrical_downtilt_min=getattr(stats, "electrical_min", None),
                electrical_downtilt_max=getattr(stats, "electrical_max", None),
            )
        )

    return items, total


def _build_lld_cells_query(
    db: Session,
    status: Optional[str],
    band_list: List[str],
    start_time: Optional[datetime],
    end_time: Optional[datetime],
):
    query = (
        db.query(SitePlanningCell, SitePlanning, Site)
        .join(SitePlanning, SitePlanningCell.planning_id == SitePlanning.id)
        .join(Site, SitePlanningCell.site_id == Site.id)
        .filter(SitePlanning.is_current == True)
    )

    if status:
        query = query.filter(Site.status == status)
    if start_time:
        query = query.filter(SitePlanning.updated_at >= start_time)
    if end_time:
        query = query.filter(SitePlanning.updated_at <= end_time)

    if band_list:
        query = query.filter(SitePlanningCell.band_code.in_(band_list))

    return query


def _fetch_lld_cells(
    db: Session,
    status: Optional[str],
    band: Optional[str],
    start_time: Optional[datetime],
    end_time: Optional[datetime],
    skip: int = 0,
    limit: Optional[int] = 50,
) -> Tuple[List[LldPlanningCellItem], int]:
    band_list = _parse_band_list(band)
    query = _build_lld_cells_query(db, status, band_list, start_time, end_time)
    total = query.count()

    if limit is not None:
        rows = (
            query.order_by(SitePlanning.updated_at.desc(), SitePlanningCell.id.desc())
            .offset(skip)
            .limit(limit)
            .all()
        )
    else:
        rows = query.order_by(SitePlanning.updated_at.desc(), SitePlanningCell.id.desc()).all()

    items: List[LldPlanningCellItem] = []
    for cell, planning, site in rows:
        base = PlanningCell.model_validate(cell, from_attributes=True).model_dump()
        base.update(
            {
                "site_code": site.site_code,
                "site_name": site.site_name,
                "site_status": site.status,
                "site_city": site.city,
                "planning_version": planning.version,
                "planning_created_at": planning.created_at,
                "planning_updated_at": planning.updated_at,
            }
        )
        items.append(LldPlanningCellItem(**base))

    return items, total


@router.get("/planning/lld-list", response_model=LldPlanningSummaryListResponse)
async def list_lld_planning(
    status: Optional[str] = Query(None, description="站点状态"),
    band: Optional[str] = Query(None, description="Band 过滤，多个用逗号分隔"),
    start_time: Optional[datetime] = Query(None, description="规划更新时间起"),
    end_time: Optional[datetime] = Query(None, description="规划更新时间止"),
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(50, ge=1, le=200, description="每页记录数"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    items, total = _fetch_lld_planning_summary(
        db=db,
        status=status,
        band=band,
        start_time=start_time,
        end_time=end_time,
        skip=skip,
        limit=limit,
    )

    active_opening_sites = _active_opening_work_orders_site_ids(db, [i.site_id for i in items])
    for it in items:
        it.edit_policy = _compute_lld_edit_policy(it, it.site_id in active_opening_sites, current_user)

    page = (skip // limit) + 1 if limit else 1
    pages = math.ceil(total / limit) if limit else 1

    return LldPlanningSummaryListResponse(
        items=items,
        total=total,
        page=page,
        size=limit,
        pages=pages,
    )


@router.get("/planning/lld-cells", response_model=LldPlanningCellListResponse)
async def list_lld_cells(
    status: Optional[str] = Query(None, description="站点状态"),
    band: Optional[str] = Query(None, description="Band 过滤，多个用逗号分隔"),
    start_time: Optional[datetime] = Query(None, description="规划更新时间起"),
    end_time: Optional[datetime] = Query(None, description="规划更新时间止"),
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(50, ge=1, le=200, description="每页记录数"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    items, total = _fetch_lld_cells(
        db=db,
        status=status,
        band=band,
        start_time=start_time,
        end_time=end_time,
        skip=skip,
        limit=limit,
    )

    page = (skip // limit) + 1 if limit else 1
    pages = math.ceil(total / limit) if limit else 1

    return LldPlanningCellListResponse(
        items=items,
        total=total,
        page=page,
        size=limit,
        pages=pages,
    )


@router.get("/planning/lld-cells/template", response_model=LldTemplateCellListResponse)
async def list_lld_cells_template(
    rat: str = Query(..., description="LTE|NR"),
    status: Optional[str] = Query(None, description="站点状态"),
    band: Optional[str] = Query(None, description="Band 过滤，多个用逗号分隔"),
    start_time: Optional[datetime] = Query(None, description="规划更新时间起"),
    end_time: Optional[datetime] = Query(None, description="规划更新时间止"),
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(50, ge=1, le=200, description="每页记录数"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    rat_norm = (rat or "").strip().upper()
    if rat_norm not in ["LTE", "NR"]:
        raise HTTPException(status_code=400, detail="rat 仅支持 LTE|NR")

    sheet = "4G" if rat_norm == "LTE" else "5G"
    headers_map = _get_lld_template_headers()
    headers = headers_map.get(sheet) or []

    band_list = _parse_band_list(band)
    query = _build_lld_cells_query(db, status, band_list, start_time, end_time).filter(SitePlanningCell.rat == rat_norm)
    # 模板视图用于回导：过滤掉无法作为有效 key 的行
    query = query.filter(SitePlanningCell.local_cell_id.isnot(None))
    query = query.filter(SitePlanningCell.band_code.isnot(None))
    query = query.filter(SitePlanningCell.band_code != "")

    total = query.count()
    rows = (
        query.order_by(SitePlanning.updated_at.desc(), SitePlanningCell.id.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    items: List[Dict[str, Any]] = []
    for cell, _planning, site in rows:
        row_data: Dict[str, Any] = {}
        for h in headers:
            row_data[h] = _to_excel_value(_get_lld_template_value(h, cell, site))
        items.append(row_data)

    page = (skip // limit) + 1 if limit else 1
    pages = math.ceil(total / limit) if limit else 1

    return LldTemplateCellListResponse(
        sheet=sheet,
        headers=headers,
        items=items,
        total=total,
        page=page,
        size=limit,
        pages=pages,
    )


@router.get("/planning/lld-list/export")
async def export_lld_planning(
    status: Optional[str] = Query(None, description="站点状态"),
    band: Optional[str] = Query(None, description="Band 过滤，多个用逗号分隔"),
    start_time: Optional[datetime] = Query(None, description="规划更新时间起"),
    end_time: Optional[datetime] = Query(None, description="规划更新时间止"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from fastapi.responses import StreamingResponse

    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    items, _ = _fetch_lld_planning_summary(
        db=db,
        status=status,
        band=band,
        start_time=start_time,
        end_time=end_time,
        skip=0,
        limit=None,
    )

    def format_range(min_val: Optional[float], max_val: Optional[float]) -> str:
        if min_val is None or max_val is None:
            return "-"
        return f"{min_val}° ~ {max_val}°"

    def format_dt(value: Optional[datetime]) -> str:
        if not value:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)

    columns = [
        ("站点编码", "site_code"),
        ("站点名称", "site_name"),
        ("类型", "site_type"),
        ("城市", "city"),
        ("状态", "status"),
        ("LLD 版本", "planning_version"),
        ("Bands", "bands"),
        ("扇区数", "sector_count"),
        ("4G Cells", "lte_cell_count"),
        ("5G Cells", "nr_cell_count"),
        ("机械下倾", "mechanical_range"),
        ("电子下倾", "electrical_range"),
        ("规划更新时间", "planning_updated_at"),
        ("规划创建时间", "planning_created_at"),
        ("备注", "planning_notes"),
    ]

    rows: List[Dict[str, Any]] = []
    for item in items:
        bands = item.bands or []
        bands_text = ", ".join(bands) if bands else "-"
        row = {
            "站点编码": item.site_code,
            "站点名称": item.site_name,
            "类型": item.site_type or "",
            "城市": item.city or "",
            "状态": item.status or "",
            "LLD 版本": item.planning_version,
            "Bands": bands_text,
            "扇区数": item.sector_count,
            "4G Cells": item.lte_cell_count,
            "5G Cells": item.nr_cell_count,
            "机械下倾": format_range(item.mechanical_downtilt_min, item.mechanical_downtilt_max),
            "电子下倾": format_range(item.electrical_downtilt_min, item.electrical_downtilt_max),
            "规划更新时间": format_dt(item.planning_updated_at),
            "规划创建时间": format_dt(item.planning_created_at),
            "备注": item.planning_notes or "",
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=[col[0] for col in columns])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="LLD汇总", index=False)
    output.seek(0)

    headers = {"Content-Disposition": "attachment; filename=lld_planning_summary.xlsx"}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/planning/lld-cells/export")
async def export_lld_cells(
    status: Optional[str] = Query(None, description="站点状态"),
    band: Optional[str] = Query(None, description="Band 过滤，多个用逗号分隔"),
    start_time: Optional[datetime] = Query(None, description="规划更新时间起"),
    end_time: Optional[datetime] = Query(None, description="规划更新时间止"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from fastapi.responses import StreamingResponse

    if current_user.role not in ["admin", "manager", "planner"]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    template_path = _get_lld_template_path()
    if not template_path.exists():
        raise HTTPException(
            status_code=500,
            detail="LLD 模板文件缺失：请在系统部署目录中放置 LLD templateV1.0.xlsx",
        )

    headers_map = _get_lld_template_headers()
    headers_4g = headers_map.get("4G") or []
    headers_5g = headers_map.get("5G") or []

    wb = load_workbook(template_path)
    if "4G" not in wb.sheetnames or "5G" not in wb.sheetnames:
        raise HTTPException(status_code=500, detail="LLD 模板缺少 4G/5G Sheet")
    ws_4g = wb["4G"]
    ws_5g = wb["5G"]

    band_list = _parse_band_list(band)
    query = _build_lld_cells_query(db, status, band_list, start_time, end_time)
    # 导出用于回导：过滤掉无法作为有效 key 的行
    query = query.filter(SitePlanningCell.local_cell_id.isnot(None))
    query = query.filter(SitePlanningCell.band_code.isnot(None))
    query = query.filter(SitePlanningCell.band_code != "")

    rows = query.order_by(SitePlanning.updated_at.desc(), SitePlanningCell.id.desc()).all()

    for cell, _planning, site in rows:
        rat = (cell.rat or "").strip().upper()
        if rat == "LTE":
            values = [
                _to_excel_value(_get_lld_template_value(h, cell, site))
                for h in headers_4g
            ]
            ws_4g.append(values)
        elif rat == "NR":
            values = [
                _to_excel_value(_get_lld_template_value(h, cell, site))
                for h in headers_5g
            ]
            ws_5g.append(values)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {"Content-Disposition": "attachment; filename=site_planning_lld_export.xlsx"}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
